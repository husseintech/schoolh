from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from models.student import Student

DATA_FILE = "students.xlsx"


def _get_student_row(sheet, student_id: str) -> Optional[int]:
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == student_id:
            return row
    return None


def load_students() -> List[Student]:
    if not Path(DATA_FILE).exists():
        return []
    wb = load_workbook(DATA_FILE)
    sheet = wb.active
    students = []
    for row in range(2, sheet.max_row + 1):
        sid = str(sheet.cell(row=row, column=1).value or "")
        name = str(sheet.cell(row=row, column=2).value or "")
        age = str(sheet.cell(row=row, column=3).value or "")
        grade = str(sheet.cell(row=row, column=4).value or "")
        phone = str(sheet.cell(row=row, column=5).value or "")
        students.append(Student(sid, name, age, grade, phone))
    wb.close()
    return students


def save_students(students: List[Student]):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "الطلاب"

    headers = Student.headers()
    header_font = Font(bold=True, size=12)
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, stu in enumerate(students, 2):
        for col_idx, val in enumerate(stu.to_dict().values(), 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + col)].width = 20

    wb.save(DATA_FILE)


def add_student(student: Student) -> bool:
    students = load_students()
    if any(s.student_id == student.student_id for s in students):
        print("خطأ: رقم الهوية موجود مسبقًا.")
        return False
    students.append(student)
    save_students(students)
    print("تم إضافة الطالب بنجاح.")
    return True


def delete_student(student_id: str) -> bool:
    students = load_students()
    filtered = [s for s in students if s.student_id != student_id]
    if len(filtered) == len(students):
        print("خطأ: لم يتم العثور على طالب بهذا الرقم.")
        return False
    save_students(filtered)
    print("تم حذف الطالب بنجاح.")
    return True


def update_student(student_id: str, name: str, age: str, grade: str, phone: str) -> bool:
    students = load_students()
    for s in students:
        if s.student_id == student_id:
            s.name = name
            s.age = age
            s.grade = grade
            s.phone = phone
            save_students(students)
            print("تم تعديل بيانات الطالب بنجاح.")
            return True
    print("خطأ: لم يتم العثور على طالب بهذا الرقم.")
    return False


def search_students(keyword: str) -> List[Student]:
    students = load_students()
    kw = keyword.strip().lower()
    results = [s for s in students if kw in s.name.lower() or kw in s.student_id]
    return results
