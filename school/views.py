import os, io, csv, re
from collections import defaultdict
from datetime import date, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Q
from django.conf import settings
from dotenv import set_key
from .models import Profile, Student, Note, Teacher, TeacherNote, Announcement, Agenda, StudentLeave, StudentLevel, ExamAnalysis, Message, Class, Subject, UserPermission, DEFAULT_PERMISSIONS, has_perm, can_view, LessonLink, StudentLateness, SchoolInfo, Meeting, SupervisorVisit, Notification, InspectionVisit, VisitProgram, Nomination, Certificate, PushSubscription, StudentAbsence, TeacherScheduleEntry, LoginCounter, StudentSurvey, WhatsAppGroup, IncomingLetter, OutgoingLetter, TeacherFollowup, ReciprocalVisit, NoObjection, AuditLog, StudentWarning, GuardianSummons
from .forms import (StudentForm, NoteForm, StudentEditForm, TeacherForm, TeacherEditForm,
    TeacherNoteForm, AnnouncementForm, AgendaForm, AgendaCompleteForm,
    StudentLeaveForm, StudentLevelForm, ExamAnalysisForm, MessageForm,
    ClassForm, SubjectForm, StudentSurveyForm)
from .services import send_push
from .services import send_whatsapp_message
from .arabic_sort import arabic_sort_key


def sort_students(students):
    return sorted(students, key=lambda s: arabic_sort_key(s.full_name))


def log_action(user, action, details=''):
    if not user or not user.is_authenticated:
        return
    try:
        AuditLog.objects.create(
            user=user,
            user_role=user.profile.role,
            action=action,
            details=details,
        )
    except Exception:
        pass


@login_required
def audit_log_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    role = request.GET.get('role', '')
    logs = AuditLog.objects.select_related('user')
    if role:
        logs = logs.filter(user_role=role)
    logs = logs.order_by('-created_at')[:2000]
    return render(request, 'school/audit_log.html', {
        'logs': logs,
        'role': role,
        'roles': ['admin', 'teacher', 'student', 'supervisor'],
    })


def sort_students_class_first(students):
    return sorted(students, key=lambda s: (
        arabic_sort_key(s.student_class.name) if getattr(s, 'student_class', None) else (0,),
        arabic_sort_key(s.full_name),
    ))


def sort_by_student_name(items):
    return sorted(items, key=lambda x: arabic_sort_key(x.student.full_name))

MODULE_KEYS = ['students', 'teachers', 'classes', 'subjects', 'announcements', 'agenda', 'leaves', 'levels', 'exams', 'messages', 'reports', 'settings', 'notes', 'lateness', 'meetings', 'supervisor_visits', 'inspection_visits', 'visit_program', 'absence', 'schedule', 'survey', 'certificates', 'guardians', 'nominations', 'incoming', 'outgoing', 'teacher_followup', 'reciprocal_visits', 'no_objection', 'open_learning']
ACTION_KEYS = ['view', 'add', 'edit', 'delete', 'import', 'export', 'notes', 'complete', 'send', 'whatsapp', 'accounts']
MODULE_LABELS = {
    'students': 'الطلاب',
    'teachers': 'المعلمون',
    'classes': 'الصفوف',
    'subjects': 'المواد',
    'announcements': 'الإعلانات',
    'agenda': 'الأجندة',
    'leaves': 'أذونات المغادرة',
    'levels': 'مستويات الطلاب',
    'exams': 'تحليل الامتحانات',
    'messages': 'الرسائل',
    'reports': 'التقارير',
    'settings': 'الإعدادات',
    'notes': 'الملاحظات',
    'lateness': 'تأخيرات الطلاب',
    'meetings': 'اجتماعات المعلمين',
    'supervisor_visits': 'زيارات المشرفين',
    'inspection_visits': 'الزيارات الإشرافية',
    'visit_program': 'برنامج الزيارات',
    'absence': 'غياب الطلاب',
    'schedule': 'الجدول اليومي للمعلمين',
    'survey': 'المسح الصحي والاجتماعي',
    'certificates': 'شهادات التقدير',
    'guardians': 'مربو الصفوف',
    'nominations': 'ترشيح المتفوقين',
    'incoming': 'سجل الوارد',
    'outgoing': 'سجل الصادر',
    'teacher_followup': 'متابعة المعلمين',
    'reciprocal_visits': 'الزيارات التبادلية',
    'no_objection': 'لا مانع',
    'open_learning': 'التعلم المفتوح',
}
ACTION_LABELS = {
    'view': 'عرض',
    'add': 'إضافة',
    'edit': 'تعديل',
    'delete': 'حذف',
    'import': 'استيراد',
    'export': 'تصدير',
    'notes': 'ملاحظات',
    'complete': 'إكمال',
    'send': 'إرسال',
    'whatsapp': 'واتساب',
    'accounts': 'الحسابات',
}


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            try:
                if user.profile.role == 'student':
                    LoginCounter.increment()
            except (Profile.DoesNotExist, AttributeError):
                pass
            return redirect('home')
        messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
        return redirect('home')
    return redirect('home')


def logout_view(request):
    logout(request)
    return redirect('home')


def home(request):
    from datetime import datetime, timedelta
    announcements = Announcement.objects.filter(is_active=True)[:5]
    lesson_links = LessonLink.objects.filter(is_active=True)[:10]
    now = datetime.now()
    now_ago = now - timedelta(hours=1)
    school_info = SchoolInfo.objects.first()
    return render(request, 'school/home.html', {
        'announcements': announcements,
        'lesson_links': lesson_links,
        'now': now,
        'now_ago': now_ago,
        'school_info': school_info,
    })


@login_required
def dashboard(request):
    profile = request.user.profile
    if profile.role in ('admin', 'vice_principal', 'secretary'):
        students_count = Student.objects.count()
        notes_count = Note.objects.count()
        teachers_count = Teacher.objects.count()
        pending_agenda = Agenda.objects.filter(is_completed=False).count()
        today = date.today()
        arabic_months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                         'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
        today_str = f'{today.day} {arabic_months[today.month - 1]} {today.year}'
        today_lateness = StudentLateness.objects.filter(date=today).count()
        today_absence = StudentAbsence.objects.filter(absence_date=today).count()
        recent_notes = list(Note.objects.select_related('student__student_class', 'created_by').order_by('-created_at')[:6])
        return render(request, 'school/admin_dashboard.html', {
            'students_count': students_count,
            'notes_count': notes_count,
            'teachers_count': teachers_count,
            'pending_agenda': pending_agenda,
            'today_lateness': today_lateness,
            'today_absence': today_absence,
            'today_str': today_str,
            'recent_notes': recent_notes,
            'show_quick_actions': profile.role in ('admin', 'vice_principal'),
        })
    elif profile.role == 'teacher':
        try:
            teacher = request.user.teacher_profile
            classes = teacher.classes.all()
            schedule_entries = list(TeacherScheduleEntry.objects.filter(
                teacher=teacher,
            ).select_related('subject', 'student_class'))
            sent_messages = list(Message.objects.filter(
                sender=request.user, recipient__profile__role='student',
            ).select_related('recipient__student_profile').order_by('-created_at')[:10])
            teacher_notes = list(Note.objects.filter(
                created_by=request.user,
            ).select_related('student__student_class').order_by('-created_at')[:10])
            return render(request, 'school/teacher_dashboard.html', {
                'teacher': teacher,
                'classes': classes,
                'schedule_entries': schedule_entries,
                'schedule_days': SCHEDULE_DAYS,
                'period_range': range(1, SCHEDULE_PERIODS + 1),
                'sent_messages': sent_messages,
                'teacher_notes': teacher_notes,
            })
        except Teacher.DoesNotExist:
            messages.error(request, 'لا يوجد ملف معلم مرتبط بهذا الحساب')
            return redirect('logout')
    else:
        try:
            student = request.user.student_profile
            notes = student.notes.filter(is_private=False).order_by('-created_at')
            notes.update(is_read=True)
            messages_qs = Message.objects.filter(recipient=request.user, is_read=False)
            absence_count = student.absences.count()
            leaves = student.leaves.all()[:10]
            schedule_entries = []
            if student.student_class:
                schedule_entries = list(TeacherScheduleEntry.objects.filter(
                    student_class=student.student_class,
                ).select_related('subject', 'teacher'))
            has_survey = hasattr(student, 'survey')
            whatsapp_groups = []
            if has_survey and student.student_class:
                whatsapp_groups = list(student.student_class.whatsapp_groups.all())
            return render(request, 'school/student_dashboard.html', {
                'student': student,
                'notes': notes,
                'unread_messages': messages_qs.count(),
                'absence_count': absence_count,
                'leaves': leaves,
                'warnings': student.warnings.all(),
                'summons': student.summons.all(),
                'schedule_entries': schedule_entries,
                'schedule_days': SCHEDULE_DAYS,
                'period_range': range(1, SCHEDULE_PERIODS + 1),
                'survey_status': '✓ مكتمل' if has_survey else '(لم يملأ بعد)',
                'has_survey': has_survey,
                'whatsapp_groups': whatsapp_groups,
            })
        except Student.DoesNotExist:
            messages.error(request, 'لا يوجد ملف طالب مرتبط بهذا الحساب')
            return redirect('logout')


@login_required
def student_absence_report(request):
    if request.user.profile.role != 'student':
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    student = request.user.student_profile
    days_names = ['الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
    absences = list(student.absences.all().order_by('-absence_date'))
    rows = [{'date': a.absence_date, 'day': days_names[a.absence_date.weekday()]} for a in absences]
    return render(request, 'school/student_absence_report.html', {
        'student': student,
        'rows': rows,
        'total': len(rows),
        'today': date.today(),
    })


# ─── Students ─────────────────────────────────────────────────────────────────

@login_required
def add_student(request):
    if not has_perm(request.user, 'students', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            username = form.cleaned_data['username']
            password = form.cleaned_data.get('password')
            if not password:
                password = form.cleaned_data['student_id'][-6:] if len(form.cleaned_data['student_id']) >= 6 else form.cleaned_data['student_id']
            log_action(request.user, 'إضافة طالب', f'{student.full_name} ({student.student_id}) - الصف: {student.student_class}')
            messages.success(request, f'تم إضافة الطالب بنجاح\nاسم المستخدم: {username}\nكلمة المرور: {password}')
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'school/add_student.html', {'form': form})


@login_required
def student_list(request):
    user = request.user
    profile = user.profile

    if profile.role == 'admin':
        students = sort_students(Student.objects.all().select_related('student_class'))
    elif profile.role == 'teacher':
        try:
            teacher = user.teacher_profile
            classes = teacher.classes.all().order_by('name')
            students = sort_students(Student.objects.filter(student_class__in=classes).select_related('student_class'))
        except Teacher.DoesNotExist:
            students = Student.objects.none()
            classes = Class.objects.none()
    else:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    if profile.role == 'admin':
        classes = Class.objects.all().order_by('name')
    return render(request, 'school/student_list.html', {'students': students, 'classes': classes})


@login_required
def edit_student(request, student_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = StudentEditForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            log_action(request.user, 'تعديل طالب', f'{student.full_name} ({student.student_id})')
            messages.success(request, 'تم تحديث بيانات الطالب بنجاح')
            return redirect('student_list')
    else:
        form = StudentEditForm(instance=student)
    return render(request, 'school/edit_student.html', {'form': form, 'student': student})


@login_required
def delete_student(request, student_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        user = student.user
        log_action(request.user, 'حذف طالب', f'{student.full_name} ({student.student_id})')
        student.delete()
        user.delete()
        messages.success(request, 'تم حذف الطالب بنجاح')
        return redirect('student_list')
    return render(request, 'school/delete_student.html', {'student': student})


@login_required
def student_notes(request, student_id):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    user = request.user
    if user.profile.role == 'admin':
        notes = Note.objects.filter(student=student).select_related('created_by').order_by('-created_at')
    else:
        notes = Note.objects.filter(student=student, is_private=False).select_related('created_by').order_by('-created_at')
    return render(request, 'school/student_notes.html', {'student': student, 'notes': notes})


@login_required
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    notes = Note.objects.filter(student=student, is_private=False).select_related('created_by').order_by('-created_at')
    levels = StudentLevel.objects.filter(student=student).select_related('subject', 'created_by').order_by('-created_at')
    return render(request, 'school/student_detail.html', {
        'student': student,
        'notes': notes,
        'levels': levels,
        'warnings': student.warnings.all(),
        'summons': student.summons.all(),
    })


@login_required
def student_discipline(request):
    if not has_perm(request.user, 'discipline', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    students = sort_students(Student.objects.all().select_related('student_class'))
    return render(request, 'school/student_discipline.html', {'students': students})


@login_required
def add_student_warning(request, student_id):
    if not has_perm(request.user, 'discipline', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        incident_date = request.POST.get('incident_date', '').strip()
        incident_facts = request.POST.get('incident_facts', '').strip()
        delivery_method = request.POST.get('delivery_method', 'direct')
        if not incident_date:
            messages.error(request, 'يرجى إدخال تاريخ الحادثة')
        elif not incident_facts:
            messages.error(request, 'يرجى إدخال وقائع الحادثة')
        else:
            warning = StudentWarning.objects.create(
                student=student,
                incident_date=incident_date,
                incident_facts=incident_facts,
                delivery_method=delivery_method,
                created_by=request.user,
            )
            log_action(request.user, 'توجيه إنذار', f'{student.full_name} ({student.student_id})')
            messages.success(request, f'تم توجيه إنذار للطالب {student.full_name}')
            return redirect('print_student_warning', warning_id=warning.id)
    return render(request, 'school/add_student_warning.html', {'student': student})


@login_required
def add_guardian_summons(request, student_id):
    if not has_perm(request.user, 'discipline', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        summons_date = request.POST.get('summons_date', '').strip()
        summons_text = request.POST.get('summons_text', '').strip()
        if not summons_date:
            messages.error(request, 'يرجى إدخال تاريخ الاستدعاء')
        elif not summons_text:
            messages.error(request, 'يرجى إدخال نص الاستدعاء')
        else:
            summons = GuardianSummons.objects.create(
                student=student,
                summons_date=summons_date,
                summons_text=summons_text,
                created_by=request.user,
            )
            log_action(request.user, 'استدعاء ولي أمر', f'{student.full_name} ({student.student_id})')
            messages.success(request, f'تم إصدار استدعاء ولي أمر للطالب {student.full_name}')
            return redirect('print_guardian_summons', summons_id=summons.id)
    return render(request, 'school/add_guardian_summons.html', {'student': student})


@login_required
def print_student_warning(request, warning_id):
    if not has_perm(request.user, 'discipline', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    warning = get_object_or_404(StudentWarning, id=warning_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/print_student_warning.html', {'warning': warning, 'info': info})


@login_required
def print_guardian_summons(request, summons_id):
    if not has_perm(request.user, 'discipline', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    summons = get_object_or_404(GuardianSummons, id=summons_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/print_guardian_summons.html', {'summons': summons, 'info': info})


@login_required
def print_all_warnings(request):
    if not has_perm(request.user, 'discipline', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    warnings = StudentWarning.objects.all().select_related('student__student_class', 'created_by').order_by('student__full_name', '-incident_date')
    info = SchoolInfo.objects.first()
    return render(request, 'school/print_all_warnings.html', {'warnings': warnings, 'info': info})


@login_required
def print_all_summons(request):
    if not has_perm(request.user, 'discipline', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    summons = GuardianSummons.objects.all().select_related('student__student_class', 'created_by').order_by('student__full_name', '-summons_date')
    info = SchoolInfo.objects.first()
    return render(request, 'school/print_all_summons.html', {'summons': summons, 'info': info})


@login_required
def reset_student_password(request, student_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        if len(new_password) < 4:
            messages.error(request, 'كلمة المرور يجب أن تكون 4 أحرف على الأقل')
            return redirect('reset_student_password', student_id=student.id)
        student.user.set_password(new_password)
        student.user.save()
        student.plain_password = new_password
        student.save()
        messages.success(request, f'تم تغيير كلمة مرور الطالب {student.full_name} إلى: {new_password}')
        return redirect('student_detail', student_id=student.id)
    return render(request, 'school/reset_password.html', {'student': student})


@login_required
def student_report(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    leaves = StudentLeave.objects.filter(student=student).order_by('-created_at')
    notes = Note.objects.filter(student=student).select_related('created_by').order_by('-created_at')
    warnings = student.warnings.all()
    summons = student.summons.all()
    days_names = ['الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
    absences = [{'date': a.absence_date, 'day': days_names[a.absence_date.weekday()]} for a in student.absences.all().order_by('-absence_date')]
    lateness = student.lateness.select_related('created_by').order_by('-date')
    principal_name = ''
    principal_phone = ''
    if request.user.profile.role == 'admin':
        principal_name = request.user.first_name or request.user.username
        principal_phone = request.user.profile.phone
    return render(request, 'school/student_report.html', {
        'student': student,
        'leaves': leaves,
        'notes': notes,
        'warnings': warnings,
        'summons': summons,
        'absences': absences,
        'lateness': lateness,
        'principal_name': principal_name,
        'principal_phone': principal_phone,
    })


# ─── Excel Import/Export ──────────────────────────────────────────────────────

@login_required
def download_student_template(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'نموذج استيراد الطلاب'

    headers = ['الاسم الكامل', 'رقم الهوية', 'الصف', 'هاتف ولي الأمر', 'اسم ولي الأمر', 'العنوان', 'تاريخ الميلاد', 'كلمة المرور', 'اسم المستخدم']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18

    note_cell = ws.cell(row=2, column=10, value='كلمة المرور اختيارية - اذا تركت فارغة سيتم إنشاءها تلقائياً من آخر 6 أرقام من رقم الهوية، واسم المستخدم إذا تُرك فارغاً يصبح رقم الهوية')
    note_cell.font = openpyxl.styles.Font(size=9, color="e74c3c")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=نموذج_استيراد_الطلاب.xlsx'
    wb.save(response)
    return response


@login_required
def import_students(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'الرجاء رفع ملف Excel صالح (xlsx أو xls)')
            return redirect('student_list')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            messages.error(request, 'فشل في قراءة الملف. تأكد من أنه ملف Excel صالح.')
            return redirect('student_list')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            messages.warning(request, 'الملف فارغ، لا توجد بيانات للاستيراد')
            return redirect('student_list')

        imported = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            try:
                full_name = str(row[0]).strip() if row[0] else ''
                student_id = str(row[1]).strip() if row[1] else ''
                class_name = str(row[2]).strip() if row[2] else ''
                parent_phone = str(row[3]).strip() if row[3] else ''
                parent_name = str(row[4]).strip() if row[4] else ''
                address = str(row[5]).strip() if row[5] else ''
                birth_date = row[6]
                password = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                username = str(row[8]).strip() if len(row) > 8 and row[8] else ''

                if not full_name or not student_id:
                    errors.append(f'الصف {i}: الاسم ورقم الهوية مطلوبان')
                    continue

                if Student.objects.filter(student_id=student_id).exists():
                    errors.append(f'الصف {i}: رقم الهوية {student_id} موجود مسبقاً')
                    continue

                class_obj = None
                if class_name:
                    class_obj, _ = Class.objects.get_or_create(name=class_name)

                if isinstance(birth_date, datetime):
                    bd = birth_date.date()
                elif isinstance(birth_date, date):
                    bd = birth_date
                else:
                    bd = None

                if not username:
                    username = student_id
                if not password:
                    password = student_id[-6:] if len(student_id) >= 6 else student_id

                user = User.objects.create_user(username=username, password=password)
                Profile.objects.create(user=user, role='student')
                Student.objects.create(
                    user=user,
                    student_id=student_id,
                    full_name=full_name,
                    student_class=class_obj,
                    parent_phone=parent_phone,
                    parent_name=parent_name,
                    address=address,
                    birth_date=bd,
                    plain_password=password,
                )
                imported += 1
            except Exception as e:
                errors.append(f'الصف {i}: خطأ - {str(e)}')

        if imported:
            msg = f'تم استيراد {imported} طالب/طالب بنجاح'
            log_action(request.user, 'استيراد طلاب من Excel', msg)
            messages.success(request, f'{msg}\nاسم المستخدم: رقم الهوية (أو ما أُدرج في عمود اسم المستخدم)\nكلمة المرور: آخر 6 أرقام من رقم الهوية (أو ما أُدرج في عمود كلمة المرور)')
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f'و {len(errors) - 10} خطأ آخر...')
        return redirect('student_list')

    return redirect('student_list')


@login_required
def export_students(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الطلاب'

    headers = ['الاسم الكامل', 'رقم الهوية', 'الصف', 'هاتف ولي الأمر', 'اسم ولي الأمر', 'العنوان', 'تاريخ الميلاد']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)

    students = sort_students_class_first(Student.objects.all().select_related('student_class'))
    for s in students:
        ws.append([
            s.full_name,
            s.student_id,
            s.student_class.name if s.student_class else '',
            s.parent_phone,
            s.parent_name,
            s.address,
            s.birth_date.strftime('%Y-%m-%d') if s.birth_date else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=قائمة_الطلاب.xlsx'
    wb.save(response)
    return response


# ─── Notes ────────────────────────────────────────────────────────────────────

@login_required
def add_note(request, student_id=None):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.created_by = request.user
            note.save()
            student = note.student
            if student.parent_phone:
                note_type = note.get_note_type_display()
                msg = f'السلام عليكم، طالبكم {student.full_name}\nنوع الملاحظة: {note_type}\nالتفاصيل: {note.content}'
                sent = send_whatsapp_message(student.parent_phone, msg)
                if sent:
                    messages.success(request, 'تم إضافة الملاحظة بنجاح وتم إرسال إشعار واتساب لولي الأمر')
                else:
                    messages.warning(request, 'تم إضافة الملاحظة ولكن تعذر إرسال رسالة واتساب')
            else:
                messages.success(request, 'تم إضافة الملاحظة بنجاح (لا يوجد رقم هاتف مسجل لولي الأمر)')
            return redirect('note_list')
    else:
        initial = {}
        if student_id:
            initial['student'] = get_object_or_404(Student, id=student_id)
        form = NoteForm(initial=initial)
        if request.user.profile.role == 'teacher':
            try:
                teacher = request.user.teacher_profile
                form.fields['student'].queryset = Student.objects.filter(student_class__in=teacher.classes.all())
            except Teacher.DoesNotExist:
                form.fields['student'].queryset = Student.objects.none()
    return render(request, 'school/add_note.html', {'form': form})


@login_required
def note_list(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.user.profile.role == 'admin':
        notes = Note.objects.all().select_related('student', 'created_by').order_by('-created_at')
    else:
        try:
            teacher = request.user.teacher_profile
            notes = Note.objects.filter(
                created_by=request.user
            ).select_related('student').order_by('-created_at')
        except Teacher.DoesNotExist:
            notes = Note.objects.none()
    return render(request, 'school/note_list.html', {'notes': notes})


# ─── Teachers ─────────────────────────────────────────────────────────────────

@login_required
def add_teacher(request):
    if not has_perm(request.user, 'teachers', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            teacher = form.save()
            log_action(request.user, 'إضافة معلم', f'{teacher.full_name} ({teacher.id_number})')
            messages.success(request, 'تم إضافة المعلم بنجاح')
            return redirect('teacher_list')
    else:
        form = TeacherForm()
    return render(request, 'school/add_teacher.html', {'form': form})


@login_required
def teacher_list(request):
    if not has_perm(request.user, 'teachers', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().prefetch_related('classes', 'subjects').order_by('full_name')
    return render(request, 'school/teacher_list.html', {'teachers': teachers})


@login_required
def download_teacher_template(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'نموذج استيراد المعلمين'

    headers = ['الاسم الكامل', 'رقم الهوية', 'البريد الإلكتروني', 'رقم الهاتف', 'المؤهل العلمي', 'التخصص', 'تاريخ الميلاد', 'تاريخ التعيين', 'كلمة المرور', 'اسم المستخدم']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    widths = [30, 20, 25, 20, 15, 20, 15, 15, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    note_cell = ws.cell(row=2, column=12, value='رقم الهوية إلزامي - اسم المستخدم إذا تُرك فارغاً يصبح رقم الهوية، وكلمة المرور إذا تُركت فارغة تصبح آخر 6 أرقام من رقم الهوية. لا تُدرج الصفوف أو المواد هنا: تُضاف لاحقاً من صفحة تعديل المعلم داخل الموقع.')
    note_cell.font = openpyxl.styles.Font(size=9, color="e74c3c")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=نموذج_استيراد_المعلمين.xlsx'
    wb.save(response)
    return response


@login_required
def import_teachers(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('teacher_list')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'الرجاء رفع ملف Excel صالح (xlsx أو xls)')
            return redirect('teacher_list')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            messages.error(request, 'فشل في قراءة الملف. تأكد من أنه ملف Excel صالح.')
            return redirect('teacher_list')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            messages.warning(request, 'الملف فارغ، لا توجد بيانات للاستيراد')
            return redirect('teacher_list')

        imported = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            try:
                full_name = str(row[0]).strip() if row[0] else ''
                id_number = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                phone = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                qualification = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                specialization = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                birth_date = row[6] if len(row) > 6 else None
                hire_date = row[7] if len(row) > 7 else None
                password = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                username = str(row[9]).strip() if len(row) > 9 and row[9] else ''

                if not full_name or not id_number:
                    errors.append(f'الصف {i}: الاسم ورقم الهوية مطلوبان')
                    continue

                if Teacher.objects.filter(id_number=id_number).exists():
                    errors.append(f'الصف {i}: رقم الهوية {id_number} موجود مسبقاً')
                    continue

                if not username:
                    username = id_number
                if User.objects.filter(username=username).exists():
                    errors.append(f'الصف {i}: اسم المستخدم {username} مستخدم مسبقاً')
                    continue
                if not password:
                    password = id_number[-6:] if len(id_number) >= 6 else id_number

                if isinstance(birth_date, datetime):
                    bd = birth_date.date()
                elif isinstance(birth_date, date):
                    bd = birth_date
                else:
                    bd = None

                if isinstance(hire_date, datetime):
                    hd = hire_date.date()
                elif isinstance(hire_date, date):
                    hd = hire_date
                else:
                    hd = None

                user = User.objects.create_user(username=username, password=password)
                Profile.objects.create(user=user, role='teacher')
                Teacher.objects.create(
                    user=user,
                    full_name=full_name,
                    id_number=id_number,
                    email=email,
                    phone=phone,
                    qualification=qualification,
                    specialization=specialization,
                    birth_date=bd,
                    hire_date=hd,
                    plain_password=password,
                )
                imported += 1
            except Exception as e:
                errors.append(f'الصف {i}: خطأ - {str(e)}')

        if imported:
            log_action(request.user, 'استيراد معلمين من Excel', f'تم استيراد {imported} معلم')
            messages.success(request, f'تم استيراد {imported} معلم بنجاح\nاسم المستخدم: رقم الهوية (أو ما أُدرج في عمود اسم المستخدم)\nكلمة المرور: آخر 6 أرقام من رقم الهوية (أو ما أُدرج في عمود كلمة المرور)')
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f'و {len(errors) - 10} خطأ آخر...')
        return redirect('teacher_list')

    return redirect('teacher_list')


@login_required
def edit_teacher(request, teacher_id):
    if not has_perm(request.user, 'teachers', 'edit'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        form = TeacherEditForm(request.POST, instance=teacher)
        if form.is_valid():
            teacher = form.save()
            password = form.cleaned_data.get('password')
            if password and teacher.user:
                teacher.user.set_password(password)
                teacher.user.save()
                teacher.plain_password = password
                teacher.save()
            log_action(request.user, 'تعديل معلم', f'{teacher.full_name} ({teacher.id_number})')
            messages.success(request, 'تم تحديث بيانات المعلم بنجاح')
            return redirect('teacher_list')
    else:
        form = TeacherEditForm(instance=teacher)
    return render(request, 'school/edit_teacher.html', {'form': form, 'teacher': teacher})


@login_required
def delete_teacher(request, teacher_id):
    if not has_perm(request.user, 'teachers', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        user = teacher.user
        log_action(request.user, 'حذف معلم', f'{teacher.full_name} ({teacher.id_number})')
        teacher.delete()
        user.delete()
        messages.success(request, 'تم حذف المعلم بنجاح')
        return redirect('teacher_list')
    return render(request, 'school/delete_teacher.html', {'teacher': teacher})


@login_required
def teacher_notes(request, teacher_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    notes = TeacherNote.objects.filter(teacher=teacher).select_related('created_by').order_by('-created_at')
    return render(request, 'school/teacher_notes.html', {'teacher': teacher, 'notes': notes})


@login_required
def teacher_detail(request, teacher_id):
    if not has_perm(request.user, 'teachers', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    notes = TeacherNote.objects.filter(teacher=teacher).select_related('created_by').order_by('-created_at')
    followups = TeacherFollowup.objects.filter(teacher=teacher).order_by('-follow_date')
    return render(request, 'school/teacher_detail.html', {
        'teacher': teacher,
        'notes': notes,
        'followups': followups,
    })


@login_required
def reset_teacher_password(request, teacher_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        if len(new_password) < 4:
            messages.error(request, 'كلمة المرور يجب أن تكون 4 أحرف على الأقل')
            return redirect('reset_teacher_password', teacher_id=teacher.id)
        teacher.user.set_password(new_password)
        teacher.user.save()
        teacher.plain_password = new_password
        teacher.save()
        log_action(request.user, 'تغيير كلمة مرور معلم', f'{teacher.full_name} ({teacher.id_number})')
        messages.success(request, f'تم تغيير كلمة مرور المعلم {teacher.full_name} إلى: {new_password}')
        return redirect('teacher_detail', teacher_id=teacher.id)
    return render(request, 'school/reset_teacher_password.html', {'teacher': teacher})


@login_required
def teachers_report(request):
    if not has_perm(request.user, 'teachers', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().prefetch_related('classes', 'subjects').order_by('full_name')
    return render(request, 'school/teachers_report.html', {'teachers': teachers})


@login_required
def teacher_cards_report(request):
    if not has_perm(request.user, 'teachers', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    all_teachers = Teacher.objects.all().order_by('full_name')
    teacher_id = request.GET.get('teacher_id', '')
    selected_teacher = None
    if teacher_id:
        selected_teacher = get_object_or_404(Teacher, id=teacher_id)
        teachers = Teacher.objects.filter(id=selected_teacher.id).prefetch_related('classes', 'subjects', 'inspection_visits', 'supervisor_visits')
    else:
        teachers = all_teachers.prefetch_related('classes', 'subjects', 'inspection_visits', 'supervisor_visits')
    info = SchoolInfo.objects.first()
    return render(request, 'school/teacher_cards_report.html', {
        'teachers': teachers,
        'all_teachers': all_teachers,
        'selected_teacher': selected_teacher,
        'info': info,
    })


@login_required
def notes_report(request):
    if not has_perm(request.user, 'notes', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    notes = Note.objects.all().select_related('student__student_class', 'created_by').order_by('-created_at')
    return render(request, 'school/notes_report.html', {'notes': notes})


@login_required
def school_info_view(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    info = SchoolInfo.objects.first()
    if request.method == 'POST':
        name_ar = request.POST.get('name_ar', '')
        name_en = request.POST.get('name_en', '')
        principal_name = request.POST.get('principal_name', '')
        national_number = request.POST.get('national_number', '')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        if latitude:
            latitude = latitude.replace(',', '.').replace('،', '.').strip()
        if longitude:
            longitude = longitude.replace(',', '.').replace('،', '.').strip()
        try:
            lat_val = float(latitude) if latitude else None
            lon_val = float(longitude) if longitude else None
        except ValueError:
            messages.error(request, 'خطأ في تنسيق الإحداثيات، استخدم النقطة (.) كفاصل عشري')
            return redirect('school_info')
        if info:
            info.name_ar = name_ar
            info.name_en = name_en
            info.principal_name = principal_name
            info.national_number = national_number
            info.latitude = lat_val
            info.longitude = lon_val
            info.school_logo = request.POST.get('school_logo', '')
            info.ministry_logo = request.POST.get('ministry_logo', '')
            info.save()
        else:
            info = SchoolInfo.objects.create(
                name_ar=name_ar, name_en=name_en,
                principal_name=principal_name, national_number=national_number,
                latitude=lat_val, longitude=lon_val,
                school_logo=request.POST.get('school_logo', ''),
                ministry_logo=request.POST.get('ministry_logo', ''),
            )
        messages.success(request, 'تم حفظ بيانات المدرسة')
        return redirect('school_info')
    return render(request, 'school/school_info.html', {'info': info})


@login_required
def add_meeting(request):
    if not has_perm(request.user, 'meetings', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().order_by('full_name')
    if request.method == 'POST':
        meeting_type = request.POST.get('meeting_type')
        date_val = request.POST.get('date')
        day = request.POST.get('day')
        time_val = request.POST.get('time')
        place = request.POST.get('place')
        goals = request.POST.get('goals', '')
        minutes = request.POST.get('minutes', '')
        meeting_number = request.POST.get('meeting_number')
        all_teachers = request.POST.get('all_teachers') == 'on'
        attendee_ids = request.POST.getlist('attendees')
        meeting = Meeting.objects.create(
            meeting_type=meeting_type, date=date_val, day=day, time=time_val,
            place=place, goals=goals, minutes=minutes,
            meeting_number=meeting_number, all_teachers=all_teachers,
            created_by=request.user
        )
        if all_teachers:
            meeting.attendees.set(Teacher.objects.all())
        elif attendee_ids:
            meeting.attendees.set(attendee_ids)
        messages.success(request, f'تم تسجيل الاجتماع رقم {meeting_number}')
        return redirect('meeting_list')
    return render(request, 'school/add_meeting.html', {'teachers': teachers})


@login_required
def meeting_list(request):
    if not has_perm(request.user, 'meetings', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    meetings = Meeting.objects.all().order_by('-date')
    return render(request, 'school/meeting_list.html', {'meetings': meetings})


@login_required
def meeting_report(request, meeting_id):
    if not has_perm(request.user, 'meetings', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    meeting = get_object_or_404(Meeting.objects.prefetch_related('attendees'), id=meeting_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/meeting_report.html', {'meeting': meeting, 'info': info})


@login_required
def add_teacher_note(request, teacher_id):
    if not has_perm(request.user, 'teachers', 'notes'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        form = TeacherNoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.teacher = teacher
            note.created_by = request.user
            note.save()
            messages.success(request, 'تم إضافة الملاحظة بنجاح')
            return redirect('teacher_notes', teacher_id=teacher.id)
    else:
        form = TeacherNoteForm(initial={'teacher': teacher})
    return render(request, 'school/add_teacher_note.html', {'form': form, 'teacher': teacher})


# ─── Classes ──────────────────────────────────────────────────────────────────

@login_required
def add_class(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            cls = form.save()
            log_action(request.user, 'إضافة صف', cls.name)
            messages.success(request, 'تم إضافة الصف بنجاح')
            return redirect('class_list')
    else:
        form = ClassForm()
    return render(request, 'school/add_class.html', {'form': form})


@login_required
def class_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    return render(request, 'school/class_list.html', {'classes': classes})


@login_required
def delete_class(request, class_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    class_obj = get_object_or_404(Class, id=class_id)
    if request.method == 'POST':
        log_action(request.user, 'حذف صف', class_obj.name)
        class_obj.delete()
        messages.success(request, 'تم حذف الصف بنجاح')
        return redirect('class_list')
    return render(request, 'school/delete_class.html', {'class_obj': class_obj})


# ─── Subjects ─────────────────────────────────────────────────────────────────

@login_required
def add_subject(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            subj = form.save()
            log_action(request.user, 'إضافة مادة', subj.name)
            messages.success(request, 'تم إضافة المادة بنجاح')
            return redirect('subject_list')
    else:
        form = SubjectForm()
    return render(request, 'school/add_subject.html', {'form': form})


@login_required
def subject_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    subjects = Subject.objects.all().order_by('name')
    return render(request, 'school/subject_list.html', {'subjects': subjects})


@login_required
def delete_subject(request, subject_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    subject = get_object_or_404(Subject, id=subject_id)
    if request.method == 'POST':
        log_action(request.user, 'حذف مادة', subject.name)
        subject.delete()
        messages.success(request, 'تم حذف المادة بنجاح')
        return redirect('subject_list')
    return render(request, 'school/delete_subject.html', {'subject': subject})


# ─── Announcements ────────────────────────────────────────────────────────────

@login_required
def announcement_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    announcements = Announcement.objects.all().order_by('-created_at')
    return render(request, 'school/announcement_list.html', {'announcements': announcements})


@login_required
def add_announcement(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.created_by = request.user
            announcement.save()
            log_action(request.user, 'إضافة إعلان', announcement.title)
            messages.success(request, 'تم إضافة الإعلان بنجاح')
            return redirect('announcement_list')
    else:
        form = AnnouncementForm()
    return render(request, 'school/add_announcement.html', {'form': form})


@login_required
def delete_announcement(request, announcement_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    announcement = get_object_or_404(Announcement, id=announcement_id)
    if request.method == 'POST':
        log_action(request.user, 'حذف إعلان', announcement.title)
        announcement.delete()
        messages.success(request, 'تم حذف الإعلان بنجاح')
        return redirect('announcement_list')
    return render(request, 'school/delete_announcement.html', {'announcement': announcement})


# ─── Agenda ───────────────────────────────────────────────────────────────────

@login_required
def agenda_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    today = date.today()
    pending_items = Agenda.objects.filter(is_completed=False).order_by('due_date')
    completed_items = Agenda.objects.filter(is_completed=True).order_by('-due_date')[:20]
    return render(request, 'school/agenda_list.html', {
        'pending_items': pending_items,
        'completed_items': completed_items,
        'today': today,
    })


@login_required
def add_agenda(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = AgendaForm(request.POST)
        if form.is_valid():
            agenda = form.save(commit=False)
            agenda.created_by = request.user
            agenda.save()
            messages.success(request, 'تم إضافة مهمة جديدة للأجندة')
            return redirect('agenda_list')
    else:
        form = AgendaForm()
    return render(request, 'school/add_agenda.html', {'form': form})


@login_required
def complete_agenda(request, agenda_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    agenda = get_object_or_404(Agenda, id=agenda_id)
    agenda.is_completed = True
    agenda.save()
    messages.success(request, f'تم إكمال المهمة: {agenda.title}')
    return redirect('agenda_list')


@login_required
def uncomplete_agenda(request, agenda_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    agenda = get_object_or_404(Agenda, id=agenda_id)
    agenda.is_completed = False
    agenda.save()
    messages.success(request, f'تم إعادة فتح المهمة: {agenda.title}')
    return redirect('agenda_list')


@login_required
def delete_agenda(request, agenda_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    agenda = get_object_or_404(Agenda, id=agenda_id)
    if request.method == 'POST':
        agenda.delete()
        messages.success(request, 'تم حذف المهمة')
        return redirect('agenda_list')
    return render(request, 'school/delete_agenda.html', {'agenda': agenda})


# ─── Student Leave ────────────────────────────────────────────────────────────

@login_required
def leave_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    leaves = StudentLeave.objects.all().select_related('student__student_class').order_by('-created_at')
    student_id = request.GET.get('student_id', '')
    search_query = request.GET.get('q', '')
    if student_id:
        leaves = leaves.filter(student_id=student_id)
    elif search_query:
        leaves = leaves.filter(student__full_name__icontains=search_query)
    students = sort_students(Student.objects.all())
    return render(request, 'school/leave_list.html', {
        'leaves': leaves,
        'students': students,
        'selected_student': student_id,
        'search_query': search_query,
    })


@login_required
def add_leave(request, student_id=None):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = StudentLeaveForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.approved_by = request.user
            leave.save()
            log_action(request.user, 'تسجيل إذن مغادرة', f'{leave.student.full_name} - {leave.date}')
            messages.success(request, 'تم تسجيل إذن المغادرة بنجاح')
            return redirect('leave_list')
    else:
        initial = {}
        if student_id:
            initial['student'] = get_object_or_404(Student, id=student_id)
        form = StudentLeaveForm(initial=initial)
        if request.user.profile.role == 'teacher':
            try:
                teacher = request.user.teacher_profile
                form.fields['student'].queryset = Student.objects.filter(student_class__in=teacher.classes.all())
            except Teacher.DoesNotExist:
                form.fields['student'].queryset = Student.objects.none()
    return render(request, 'school/add_leave.html', {'form': form})


@login_required
def delete_leave(request, leave_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    leave = get_object_or_404(StudentLeave, id=leave_id)
    if request.method == 'POST':
        log_action(request.user, 'حذف إذن مغادرة', f'{leave.student.full_name} - {leave.date}')
        leave.delete()
        messages.success(request, 'تم حذف إذن المغادرة')
        return redirect('leave_list')
    return render(request, 'school/delete_leave.html', {'leave': leave})


# ─── Student Levels ───────────────────────────────────────────────────────────

@login_required
def bulk_add_student_level(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    try:
        teacher = request.user.teacher_profile
    except Teacher.DoesNotExist:
        if request.user.profile.role != 'admin':
            messages.error(request, 'يجب أن تكون معلماً لتسجيل المستويات')
            return redirect('dashboard')
        teacher = None
    if teacher:
        classes = teacher.classes.all().order_by('name')
        subjects_qs = teacher.subjects.all()
    else:
        classes = Class.objects.all().order_by('name')
        subjects_qs = Subject.objects.all()
    if request.method == 'POST':
        students_ids = request.POST.getlist('student_id')
        levels = request.POST.getlist('level')
        notes_list = request.POST.getlist('notes')
        subject_id = request.POST.get('subject')
        class_id = request.POST.get('class_id')
        subject = Subject.objects.get(id=subject_id) if subject_id else None
        count = 0
        for i, student_id in enumerate(students_ids):
            if levels[i]:
                StudentLevel.objects.create(
                    student_id=student_id,
                    subject=subject,
                    level=levels[i],
                    notes=notes_list[i] if i < len(notes_list) else '',
                    created_by=request.user
                )
                count += 1
        messages.success(request, f'تم تسجيل {count} مستوى بنجاح')
        return redirect('student_level_list')
    class_id = request.GET.get('class_id')
    selected_class = None
    students = []
    if class_id:
        selected_class = get_object_or_404(Class, id=class_id)
        if teacher and selected_class not in teacher.classes.all():
            messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الصف')
            return redirect('bulk_add_student_level')
        students = sort_students(Student.objects.filter(student_class=selected_class))
    return render(request, 'school/bulk_add_student_level.html', {
        'students': students,
        'subjects': subjects_qs,
        'classes': classes,
        'selected_class': selected_class,
    })

def add_student_level(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    try:
        teacher = request.user.teacher_profile
    except Teacher.DoesNotExist:
        if request.user.profile.role != 'admin':
            messages.error(request, 'يجب أن تكون معلماً لتسجيل المستويات')
            return redirect('dashboard')
        teacher = None
    if teacher:
        classes = teacher.classes.all().order_by('name')
        subjects_qs = teacher.subjects.all()
    else:
        classes = Class.objects.all().order_by('name')
        subjects_qs = Subject.objects.all()
    if request.method == 'POST':
        students_ids = request.POST.getlist('student_id')
        levels = request.POST.getlist('level')
        notes_list = request.POST.getlist('notes')
        subject_id = request.POST.get('subject')
        class_id = request.POST.get('class_id')
        subject = Subject.objects.get(id=subject_id) if subject_id else None
        count = 0
        for i, student_id in enumerate(students_ids):
            if levels[i]:
                StudentLevel.objects.create(
                    student_id=student_id,
                    subject=subject,
                    level=levels[i],
                    notes=notes_list[i] if i < len(notes_list) else '',
                    created_by=request.user
                )
                count += 1
        messages.success(request, f'تم تسجيل {count} مستوى بنجاح')
        return redirect('student_level_list')
    class_id = request.GET.get('class_id')
    selected_class = None
    students = []
    if class_id:
        selected_class = get_object_or_404(Class, id=class_id)
        if teacher and selected_class not in teacher.classes.all():
            messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الصف')
            return redirect('add_student_level')
        students = sort_students(Student.objects.filter(student_class=selected_class))
    return render(request, 'school/add_student_level.html', {
        'students': students,
        'subjects': subjects_qs,
        'classes': classes,
        'selected_class': selected_class,
    })


@login_required
def student_level_list(request):
    user = request.user
    is_teacher = user.profile.role == 'teacher'
    if user.profile.role == 'admin':
        levels_qs = StudentLevel.objects.all().select_related('student', 'subject', 'created_by').order_by('-created_at')
    elif is_teacher:
        try:
            teacher = user.teacher_profile
        except Teacher.DoesNotExist:
            messages.error(request, 'لا يوجد حساب معلم مرتبط بحسابك')
            return redirect('dashboard')
        levels_qs = StudentLevel.objects.none()
    else:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    students = []
    selected_class = None
    selected_subject = None
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    level_filter = request.GET.get('level')

    if is_teacher:
        classes = teacher.classes.all().order_by('name')
        subjects = teacher.subjects.all().order_by('name')

    if class_id and subject_id:
        selected_class = get_object_or_404(Class, id=class_id)
        selected_subject = get_object_or_404(Subject, id=subject_id)
        if is_teacher:
            if not teacher.classes.filter(id=selected_class.id).exists() or not teacher.subjects.filter(id=selected_subject.id).exists():
                messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الصف أو المادة')
                return redirect('student_level_list')
            students = sort_students(Student.objects.filter(student_class=selected_class))
            for s in students:
                s.current_level = StudentLevel.objects.filter(student=s, subject=selected_subject).first()
            if request.method == 'POST':
                for s in students:
                    lvl = request.POST.get(f'level_{s.id}')
                    notes = request.POST.get(f'notes_{s.id}', '') or ''
                    if lvl:
                        StudentLevel.objects.update_or_create(
                            student=s, subject=selected_subject,
                            defaults={'level': lvl, 'notes': notes, 'created_by': request.user},
                        )
                messages.success(request, f'تم حفظ مستويات {len(students)} طالب')
                return redirect(f"{request.path}?class_id={selected_class.id}&subject_id={selected_subject.id}")
        else:
            levels_qs = levels_qs.filter(student__student_class=selected_class, subject=selected_subject)
    elif class_id:
        selected_class = get_object_or_404(Class, id=class_id)
        if is_teacher:
            if not teacher.classes.filter(id=selected_class.id).exists():
                messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الصف')
                return redirect('student_level_list')
        else:
            levels_qs = levels_qs.filter(student__student_class=selected_class)

    selected_level = None
    selected_level_label = None
    if level_filter and not is_teacher:
        selected_level = level_filter
        levels_qs = levels_qs.filter(level=level_filter)
        selected_level_label = dict(StudentLevel.LEVEL_CHOICES).get(level_filter)

    return render(request, 'school/student_level_list.html', {
        'is_teacher': is_teacher,
        'levels': levels_qs,
        'students': students,
        'classes': classes,
        'subjects': subjects,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'selected_level': selected_level,
        'selected_level_label': selected_level_label,
        'level_choices': StudentLevel.LEVEL_CHOICES,
    })


# ─── Exam Analysis ────────────────────────────────────────────────────────────

@login_required
def add_exam_analysis(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        form = ExamAnalysisForm(request.POST)
        if form.is_valid():
            analysis = form.save(commit=False)
            analysis.created_by = request.user
            if analysis.total_students > 0:
                analysis.pass_percentage = round((analysis.passed_count / analysis.total_students) * 100, 2)
                analysis.fail_percentage = round((analysis.failed_count / analysis.total_students) * 100, 2)
            analysis.save()
            messages.success(request, 'تم إضافة تحليل الامتحان بنجاح')
            return redirect('exam_analysis_list')
    else:
        form = ExamAnalysisForm()
        if request.user.profile.role == 'teacher':
            try:
                teacher = request.user.teacher_profile
                form.fields['subject'].queryset = teacher.subjects.all()
                form.fields['student_class'].queryset = teacher.classes.all()
            except Teacher.DoesNotExist:
                form.fields['subject'].queryset = Subject.objects.none()
                form.fields['student_class'].queryset = Class.objects.none()
    return render(request, 'school/add_exam_analysis.html', {'form': form})


@login_required
def exam_analysis_list(request):
    if request.user.profile.role == 'admin':
        analyses = ExamAnalysis.objects.all().select_related('subject', 'student_class', 'created_by').order_by('-created_at')
    elif request.user.profile.role == 'teacher':
        try:
            teacher = request.user.teacher_profile
            analyses = ExamAnalysis.objects.filter(
                created_by=request.user
            ).select_related('subject', 'student_class').order_by('-created_at')
        except Teacher.DoesNotExist:
            analyses = ExamAnalysis.objects.none()
    else:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    return render(request, 'school/exam_analysis_list.html', {'analyses': analyses})


@login_required
def exam_analysis_report(request, analysis_id):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    analysis = get_object_or_404(ExamAnalysis, id=analysis_id)
    if request.user.profile.role == 'teacher' and analysis.created_by != request.user:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('exam_analysis_list')
    info = SchoolInfo.objects.first()
    return render(request, 'school/exam_analysis_report.html', {
        'analysis': analysis,
        'info': info,
        'today': date.today(),
    })


# ─── Messages ─────────────────────────────────────────────────────────────────

@login_required
def message_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    messages_qs = Message.objects.filter(recipient__isnull=True).order_by('-created_at')
    return render(request, 'school/message_list.html', {'messages_qs': messages_qs})


@login_required
def send_message(request, user_id=None):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        recipient_id = request.POST.get('recipient_id')
        subject = request.POST.get('subject', '').strip()
        content = request.POST.get('content', '').strip()
        if not recipient_id or not subject or not content:
            messages.error(request, 'الموضوع والرسالة مطلوبان')
            return redirect('send_message_to', user_id=recipient_id or 0)
        recipient = get_object_or_404(User, id=recipient_id)
        Message.objects.create(sender=request.user, recipient=recipient, subject=subject, content=content)
        send_push(
            recipient,
            f'رسالة جديدة من {request.user.first_name or request.user.username}',
            subject[:150],
            '/messages/',
        )
        messages.success(request, f'تم إرسال الرسالة إلى {recipient.first_name or recipient.username}')
        return redirect('send_message')
    students = sort_students(Student.objects.all().select_related('student_class'))
    teachers = Teacher.objects.all().order_by('full_name')
    admins = User.objects.filter(profile__role__in=['admin', 'vice_principal', 'secretary']).select_related('profile').order_by('first_name')
    if user_id:
        recipient = get_object_or_404(User, id=user_id)
        return render(request, 'school/send_message_form.html', {'recipient': recipient})
    return render(request, 'school/send_message.html', {
        'students': students,
        'teachers': teachers,
        'admins': admins,
    })


@login_required
def messages_report(request):
    if not has_perm(request.user, 'messages', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    msgs = Message.objects.filter(recipient__profile__role='student').select_related('sender', 'recipient').order_by('-created_at')
    info = SchoolInfo.objects.first()
    return render(request, 'school/messages_report.html', {
        'messages_qs': msgs,
        'info': info,
    })


@login_required
def sent_messages(request):
    if request.user.profile.role not in ['admin', 'teacher']:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    msg_list = Message.objects.filter(sender=request.user).order_by('-created_at')
    return render(request, 'school/sent_messages.html', {'messages_qs': msg_list})


@login_required
def read_message(request, message_id):
    msg = get_object_or_404(Message, id=message_id)
    if msg.recipient is None and msg.sender is None:
        if request.user.profile.role != 'admin':
            messages.error(request, 'لا يمكنك قراءة هذه الرسالة')
            return redirect('dashboard')
    elif msg.sender != request.user and msg.recipient != request.user:
        messages.error(request, 'لا يمكنك قراءة هذه الرسالة')
        return redirect('dashboard')
    msg.is_read = True
    msg.save()
    return render(request, 'school/read_message.html', {'msg': msg})


@login_required
def student_messages(request):
    if request.user.profile.role != 'student':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    msgs = Message.objects.filter(recipient=request.user).order_by('-created_at')
    return render(request, 'school/student_messages.html', {'messages_qs': msgs})


# ─── Message Deletion ──────────────────────────────────────────────────────────

@login_required
def delete_message(request, message_id):
    msg = get_object_or_404(Message, id=message_id)
    if msg.sender != request.user and msg.recipient != request.user:
        messages.error(request, 'لا يمكنك حذف هذه الرسالة')
        return redirect('sent_messages')
    link = f'/messages/{msg.id}/read/'
    Notification.objects.filter(link=link).delete()
    msg.delete()
    messages.success(request, 'تم حذف الرسالة بنجاح')
    if request.user.profile.role == 'student':
        return redirect('student_messages')
    return redirect('sent_messages')


@login_required
def delete_all_sent_messages(request):
    count = Message.objects.filter(sender=request.user).count()
    Message.objects.filter(sender=request.user).delete()
    Notification.objects.filter(user=request.user, link__startswith='/messages/').delete()
    messages.success(request, f'تم حذف {count} رسالة مرسلة')
    return redirect('sent_messages')


@login_required
def delete_all_received_messages(request):
    count = Message.objects.filter(recipient=request.user).count()
    Message.objects.filter(recipient=request.user).delete()
    Notification.objects.filter(user=request.user, link__startswith='/messages/').delete()
    messages.success(request, f'تم حذف {count} رسالة مستلمة')
    if request.user.profile.role == 'student':
        return redirect('student_messages')
    return redirect('sent_messages')


@login_required
def delete_all_messages(request):
    sent = Message.objects.filter(sender=request.user).count()
    received = Message.objects.filter(recipient=request.user).count()
    Message.objects.filter(sender=request.user).delete()
    Message.objects.filter(recipient=request.user).delete()
    Notification.objects.filter(user=request.user, link__startswith='/messages/').delete()
    messages.success(request, f'تم مسح جميع الرسائل: {sent} مرسلة و {received} مستلمة')
    if request.user.profile.role == 'student':
        return redirect('student_messages')
    return redirect('sent_messages')


# ─── Reports ──────────────────────────────────────────────────────────────────

@login_required
def reports(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    if request.method == 'POST' and request.POST.get('action') == 'reset_login_counter':
        LoginCounter.reset()
        messages.success(request, 'تم تصفير عداد الدخول')
        return redirect('reports')
    login_counter = LoginCounter.get()
    return render(request, 'school/reports.html', {
        'classes': classes,
        'login_counter': login_counter,
    })


@login_required
def class_report(request, class_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    class_obj = get_object_or_404(Class, id=class_id)
    students = sort_students(Student.objects.filter(student_class=class_obj))
    notes = Note.objects.filter(student__student_class=class_obj).order_by('-created_at')
    today = date.today()
    absent_today_ids = list(StudentAbsence.objects.filter(
        absence_date=today, student__student_class=class_obj
    ).values_list('student_id', flat=True))

    context = {
        'class_obj': class_obj,
        'students': students,
        'notes_count': notes.count(),
        'today': today,
        'absent_today_ids': absent_today_ids,
}
    return render(request, 'school/class_report.html', context)


@login_required
def student_levels_report(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    selected_class = None
    selected_subject = None
    levels = []
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    if class_id and subject_id:
        selected_class = get_object_or_404(Class, id=class_id)
        selected_subject = get_object_or_404(Subject, id=subject_id)
        levels = StudentLevel.objects.filter(
            student__student_class=selected_class,
            subject=selected_subject
        ).select_related('student', 'created_by')
        levels = sort_by_student_name(levels)
    return render(request, 'school/student_levels_report.html', {
        'classes': classes,
        'subjects': subjects,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'levels': levels,
    })


@login_required
def reports_overview(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    total_students = Student.objects.count()
    total_teachers = Teacher.objects.count()
    total_classes = Class.objects.count()
    total_notes = Note.objects.count()
    total_leaves = StudentLeave.objects.count()
    total_exams = ExamAnalysis.objects.count()

    class_stats = []
    for cls in Class.objects.all().order_by('name'):
        student_count = Student.objects.filter(student_class=cls).count()
        notes_count = Note.objects.filter(student__student_class=cls).count()
        class_stats.append({
            'id': cls.id,
            'name': cls.name,
            'students': student_count,
            'notes': notes_count,
        })

    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_classes': total_classes,
        'total_notes': total_notes,
        'total_leaves': total_leaves,
        'total_exams': total_exams,
        'class_stats': class_stats,
    }
    return render(request, 'school/reports_overview.html', context)


# ─── Account Management ───────────────────────────────────────────────────────

@login_required
def account_list(request):
    if not has_perm(request.user, 'settings', 'accounts'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    admins = User.objects.filter(is_superuser=False, profile__role='admin').select_related('profile')
    vice_principals = User.objects.filter(is_superuser=False, profile__role='vice_principal').select_related('profile')
    secretaries = User.objects.filter(is_superuser=False, profile__role='secretary').select_related('profile')
    teachers = User.objects.filter(is_superuser=False, profile__role='teacher').select_related('profile')
    students = User.objects.filter(is_superuser=False, profile__role='student').select_related('profile')
    return render(request, 'school/account_list.html', {
        'admins': admins,
        'vice_principals': vice_principals,
        'secretaries': secretaries,
        'teachers': teachers,
        'students': students,
    })


@login_required
def add_account(request):
    if not has_perm(request.user, 'settings', 'accounts'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        full_name = request.POST.get('full_name', '').strip()
        role = request.POST.get('role', '')
        phone = request.POST.get('phone', '').strip()

        if not username or not password:
            messages.error(request, 'اسم المستخدم وكلمة المرور مطلوبان')
            return redirect('add_account')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'اسم المستخدم موجود مسبقاً')
            return redirect('add_account')

        user = User.objects.create_user(username=username, password=password, first_name=full_name)
        Profile.objects.create(user=user, role=role, phone=phone)
        if role == 'teacher':
            Teacher.objects.create(user=user, full_name=full_name, phone=phone)

        new_perms = {}
        if role == 'admin':
            new_perms = {m: list(ACTION_KEYS) for m in MODULE_KEYS}
        else:
            for key, val in request.POST.items():
                if key.startswith('perm_'):
                    parts = key.replace('perm_', '', 1).rsplit('_', 1)
                    if len(parts) == 2:
                        module, action = parts
                        new_perms.setdefault(module, []).append(action)
        if not new_perms:
            new_perms = UserPermission.get_defaults(role)
        UserPermission.objects.create(user=user, permissions=new_perms)

        log_action(request.user, 'إضافة حساب', f'{username} - {dict(Profile.ROLE_CHOICES).get(role, "")}')
        messages.success(request, f'تم إضافة الحساب: {username} - {dict(Profile.ROLE_CHOICES).get(role, "")}')
        return redirect('account_list')

    modules = [{'key': k, 'label': MODULE_LABELS[k]} for k in MODULE_KEYS]
    actions = [{'key': k, 'label': ACTION_LABELS[k]} for k in ACTION_KEYS]
    return render(request, 'school/add_account.html', {
        'roles': Profile.ROLE_CHOICES,
        'modules': modules,
        'actions': actions,
        'default_perms': DEFAULT_PERMISSIONS,
    })


@login_required
def edit_account(request, user_id):
    if not has_perm(request.user, 'settings', 'accounts'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    user = get_object_or_404(User, id=user_id)
    profile = user.profile
    perms, _ = UserPermission.objects.get_or_create(user=user, defaults={'permissions': UserPermission.get_defaults(profile.role)})

    if request.method == 'POST':
        profile.role = request.POST.get('role', profile.role)
        profile.phone = request.POST.get('phone', '')
        user.first_name = request.POST.get('full_name', '')
        profile.save()
        user.save()
        if profile.role == 'teacher':
            try:
                user.teacher_profile
            except Teacher.DoesNotExist:
                Teacher.objects.create(user=user, full_name=user.first_name or user.username, phone=profile.phone)

        new_password = request.POST.get('new_password', '').strip()
        if new_password:
            user.set_password(new_password)
            user.save()

        new_perms = {}
        if profile.role == 'admin':
            new_perms = {m: list(ACTION_KEYS) for m in MODULE_KEYS}
        else:
            for key, val in request.POST.items():
                if key.startswith('perm_'):
                    parts = key.replace('perm_', '', 1).rsplit('_', 1)
                    if len(parts) == 2:
                        module, action = parts
                        new_perms.setdefault(module, []).append(action)
        perms.permissions = new_perms
        perms.save()

        log_action(request.user, 'تعديل حساب', user.username)
        messages.success(request, f'تم تحديث الحساب: {user.username}')
        return redirect('account_list')

    modules = [{'key': k, 'label': MODULE_LABELS[k]} for k in MODULE_KEYS]
    actions = [{'key': k, 'label': ACTION_LABELS[k]} for k in ACTION_KEYS]

    allowed_set = set()
    for module, actions_list in perms.permissions.items():
        for action in actions_list:
            allowed_set.add(f'{module}_{action}')

    return render(request, 'school/edit_account.html', {
        'edit_user': user,
        'profile': profile,
        'allowed_set': allowed_set,
        'roles': Profile.ROLE_CHOICES,
        'modules': modules,
        'actions': actions,
        'default_perms': DEFAULT_PERMISSIONS,
    })


@login_required
def delete_account(request, user_id):
    if not has_perm(request.user, 'settings', 'accounts'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, 'لا يمكنك حذف حسابك الخاص')
        return redirect('account_list')
    if request.method == 'POST':
        username = user.username
        log_action(request.user, 'حذف حساب', username)
        user.delete()
        messages.success(request, f'تم حذف الحساب: {username}')
        return redirect('account_list')
    return render(request, 'school/delete_account.html', {'del_user': user})


@login_required
def role_permissions(request):
    if not has_perm(request.user, 'settings', 'accounts'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    modules = [{'key': k, 'label': MODULE_LABELS[k]} for k in MODULE_KEYS]
    actions = [{'key': k, 'label': ACTION_LABELS[k]} for k in ACTION_KEYS]

    role = request.GET.get('role', 'student')
    if role not in dict(Profile.ROLE_CHOICES):
        role = 'student'

    if request.method == 'POST':
        role = request.POST.get('role', role)
        action_kind = request.POST.get('action_kind', 'default')  # 'default' | 'custom'
        new_perms = {}
        if action_kind == 'custom':
            for key, val in request.POST.items():
                if key.startswith('perm_'):
                    parts = key.replace('perm_', '', 1).rsplit('_', 1)
                    if len(parts) == 2:
                        module, act = parts
                        new_perms.setdefault(module, []).append(act)
        else:
            new_perms = UserPermission.get_defaults(role)

        if not new_perms:
            new_perms = UserPermission.get_defaults(role)

        # ensure every user of that role has a UserPermission row
        users = User.objects.filter(profile__role=role)
        updated = 0
        for u in users:
            up, _ = UserPermission.objects.get_or_create(user=u, defaults={'permissions': new_perms})
            up.permissions = new_perms
            up.save()
            updated += 1
        role_label = dict(Profile.ROLE_CHOICES).get(role, role)
        log_action(request.user, 'تعديل صلاحيات دور', f'{role_label} ({action_kind}) - {updated} حساب')
        messages.success(request, f'تم منح الصلاحيات ({action_kind}) لـ {updated} من حسابات دور {role_label}')
        return redirect(f"{reverse('role_permissions')}?role={role}")

    counts = User.objects.filter(profile__role__in=[r for r, _ in Profile.ROLE_CHOICES]).values('profile__role').annotate(n=Count('id'))
    count_map = {c['profile__role']: c['n'] for c in counts}
    role_counts = [(r, label, count_map.get(r, 0)) for r, label in Profile.ROLE_CHOICES]
    return render(request, 'school/role_permissions.html', {
        'roles': Profile.ROLE_CHOICES,
        'modules': modules,
        'actions': actions,
        'default_perms': DEFAULT_PERMISSIONS,
        'selected_role': role,
        'role_counts': role_counts,
    })


@login_required
def my_account(request):
    user = request.user
    profile = user.profile
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        if full_name:
            user.first_name = full_name
            try:
                if hasattr(user, 'teacher_profile'):
                    t = user.teacher_profile
                    t.full_name = full_name
                    t.save()
            except Exception:
                pass
            try:
                if hasattr(user, 'student_profile'):
                    s = user.student_profile
                    s.full_name = full_name
                    s.save()
            except Exception:
                pass
        if phone:
            profile.phone = phone
        if new_password:
            user.set_password(new_password)
        profile.save()
        user.save()
        if new_password:
            messages.success(request, 'تم تغيير كلمة المرور. سجل دخول مرة أخرى')
            return redirect('login')
        messages.success(request, 'تم تحديث البيانات')
        return redirect('my_account')
    return render(request, 'school/my_account.html', {'profile': profile})


# ─── Lesson Links (Online Classes) ──────────────────────────────────────────

@login_required
def lesson_link_list(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    links = LessonLink.objects.all().order_by('-lesson_datetime', '-created_at')
    return render(request, 'school/lesson_link_list.html', {'links': links})


@login_required
def add_lesson_link(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        url = request.POST.get('url', '').strip()
        lesson_date = request.POST.get('lesson_date', '').strip()
        lesson_time = request.POST.get('lesson_time', '').strip()
        if not title or not url:
            messages.error(request, 'العنوان والرابط مطلوبان')
            return redirect('add_lesson_link')
        dt = None
        if lesson_date and lesson_time:
            from datetime import datetime as dt_mod
            try:
                dt = dt_mod.strptime(f'{lesson_date} {lesson_time}', '%Y-%m-%d %H:%M')
            except ValueError:
                pass
        LessonLink.objects.create(title=title, url=url, lesson_datetime=dt)
        messages.success(request, f'تم إضافة الرابط: {title}')
        return redirect('lesson_link_list')
    return render(request, 'school/add_lesson_link.html')


@login_required
def delete_lesson_link(request, link_id):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    link = get_object_or_404(LessonLink, id=link_id)
    if request.method == 'POST':
        title = link.title
        link.delete()
        messages.success(request, f'تم حذف الرابط: {title}')
        return redirect('lesson_link_list')
    return render(request, 'school/delete_lesson_link.html', {'link': link})


# ─── WhatsApp ─────────────────────────────────────────────────────────────────

@login_required
def whatsapp_settings(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    env_path = settings.BASE_DIR / '.env'
    if request.method == 'POST':
        provider = request.POST.get('provider', 'log')
        token = request.POST.get('ultramsg_token', '')
        instance_id = request.POST.get('ultramsg_instance_id', '')
        set_key(str(env_path), 'WHATSAPP_PROVIDER', provider)
        set_key(str(env_path), 'ULTRAMSG_TOKEN', token)
        set_key(str(env_path), 'ULTRAMSG_INSTANCE_ID', instance_id)
        os.environ['WHATSAPP_PROVIDER'] = provider
        os.environ['ULTRAMSG_TOKEN'] = token
        os.environ['ULTRAMSG_INSTANCE_ID'] = instance_id
        messages.success(request, 'تم حفظ إعدادات واتساب بنجاح')
        return redirect('dashboard')
    context = {
        'current_provider': os.getenv('WHATSAPP_PROVIDER', 'log'),
        'ultramsg_token': os.getenv('ULTRAMSG_TOKEN', ''),
        'ultramsg_instance_id': os.getenv('ULTRAMSG_INSTANCE_ID', ''),
    }
    return render(request, 'school/whatsapp_settings.html', context)


# ─── Student Lateness ─────────────────────────────────────────────────────────

@login_required
def lateness_list(request):
    if not has_perm(request.user, 'lateness', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    students = Student.objects.all().select_related('student_class')
    search_query = request.GET.get('q', '')
    if search_query:
        students = students.filter(full_name__icontains=search_query)
    students = students.annotate(lateness_count=Count('lateness'))
    students = sort_students_class_first(students)
    today_lateness = list(StudentLateness.objects.filter(date=date.today()).values_list('student_id', flat=True))
    today_notes = {
        l.student_id: l.notes
        for l in StudentLateness.objects.filter(date=date.today())
    }
    for s in students:
        s.lateness_notes = today_notes.get(s.id, '')
    if request.method == 'POST':
        if not has_perm(request.user, 'lateness', 'add'):
            messages.error(request, 'ليس لديك صلاحية للإضافة')
            return redirect('lateness_list')
        student_ids = request.POST.getlist('student_id')
        lateness_date = request.POST.get('lateness_date', str(date.today()))
        count = 0
        for sid in student_ids:
            note = request.POST.get(f'notes_{sid}', '')
            StudentLateness.objects.update_or_create(
                student_id=sid,
                date=lateness_date,
                defaults={'notes': note, 'created_by': request.user}
            )
            count += 1
        messages.success(request, f'تم تسجيل تأخير {count} طالب/طلاب')
        return redirect('lateness_list')
    return render(request, 'school/lateness_list.html', {
        'students': students,
        'classes': classes,
        'search_query': search_query,
        'today_lateness': today_lateness,
        'today': date.today(),
    })


@login_required
def lateness_report(request):
    if not has_perm(request.user, 'lateness', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    current_year = date.today().year
    month = request.GET.get('month', str(date.today().month))
    lateness_qs = StudentLateness.objects.filter(date__year=current_year).select_related('student', 'created_by')
    if month:
        lateness_qs = lateness_qs.filter(date__month=month)
    lateness_qs = lateness_qs.order_by('-date', 'student__full_name')
    months = [
        ('1', 'يناير'), ('2', 'فبراير'), ('3', 'مارس'), ('4', 'إبريل'),
        ('5', 'مايو'), ('6', 'يونيو'), ('7', 'يوليو'), ('8', 'أغسطس'),
        ('9', 'سبتمبر'), ('10', 'أكتوبر'), ('11', 'نوفمبر'), ('12', 'ديسمبر'),
    ]
    return render(request, 'school/lateness_report.html', {
        'lateness_list': lateness_qs,
        'months': months,
        'selected_month': month,
        'current_year': current_year,
    })


@login_required
def student_lateness_detail(request, student_id):
    if not has_perm(request.user, 'lateness', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    lateness_records = StudentLateness.objects.filter(student=student).select_related('created_by').order_by('-date')
    return render(request, 'school/student_lateness_detail.html', {
        'student': student,
        'lateness_records': lateness_records,
    })


# ─── Supervisor Visits ─────────────────────────────────────────────────────────

@login_required
def supervisor_visit_list(request):
    if not has_perm(request.user, 'supervisor_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().order_by('full_name')
    selected_teacher = None
    visits = SupervisorVisit.objects.none()
    teacher_id = request.GET.get('teacher_id', '')
    if teacher_id:
        selected_teacher = get_object_or_404(Teacher, id=teacher_id)
        visits = SupervisorVisit.objects.filter(teacher=selected_teacher).order_by('-visit_date')
    if request.method == 'POST' and has_perm(request.user, 'supervisor_visits', 'add'):
        teacher_id = request.POST.get('teacher_id')
        selected_teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None
        if selected_teacher:
            visit = SupervisorVisit.objects.create(
                teacher=selected_teacher,
                visit_number=request.POST.get('visit_number', ''),
                visit_date=request.POST.get('visit_date', ''),
                subject_area=request.POST.get('subject_area', ''),
                lesson_topic=request.POST.get('lesson_topic', ''),
                class_name=request.POST.get('class_name', ''),
                section=request.POST.get('section', ''),
                supervisor_name=request.POST.get('supervisor_name', ''),
                recommendations=request.POST.get('recommendations', ''),
                admin_followup=request.POST.get('admin_followup', ''),
                created_by=request.user,
            )
            # Send notification to the teacher
            if selected_teacher.user:
                recs = request.POST.get('recommendations', '').strip()
                msg = f'تم تسجيل زيارة مشرف بتاريخ {visit.visit_date}'
                if recs:
                    msg += f'\nتوصيات المشرف: {recs[:500]}'
                Notification.objects.create(
                    user=selected_teacher.user,
                    title='زيارة مشرف جديدة',
                    message=msg,
                    link=f'/supervisor-visits/{visit.id}/report/',
                )
                send_push(selected_teacher.user, 'زيارة مشرف جديدة', f'تم تسجيل زيارة مشرف بتاريخ {visit.visit_date}', f'/supervisor-visits/{visit.id}/report/')
            messages.success(request, 'تم إضافة الزيارة بنجاح')
            return redirect(f'{request.path}?teacher_id={teacher_id}')
        messages.error(request, 'الرجاء اختيار معلم')
    return render(request, 'school/supervisor_visit_list.html', {
        'teachers': teachers,
        'selected_teacher': selected_teacher,
        'visits': visits,
    })


@login_required
def supervisor_visit_report(request, visit_id):
    if not has_perm(request.user, 'supervisor_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(SupervisorVisit, id=visit_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/supervisor_visit_report.html', {
        'visit': visit,
        'info': info,
    })


@login_required
def supervisor_visits_report(request):
    if not has_perm(request.user, 'supervisor_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visits = SupervisorVisit.objects.all().select_related('teacher').order_by('-visit_date')
    teachers = Teacher.objects.all().order_by('full_name')
    selected_teacher_id = request.GET.get('teacher_id', '')
    if selected_teacher_id:
        visits = visits.filter(teacher_id=selected_teacher_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/supervisor_visits_report.html', {
        'visits': visits,
        'info': info,
        'teachers': teachers,
        'selected_teacher_id': selected_teacher_id,
    })


# ─── Visit Program ─────────────────────────────────────────────────────────────

@login_required
def visit_program_entry(request, entry_id):
    entry = get_object_or_404(VisitProgram, id=entry_id)
    if request.user.profile.role == 'teacher' and entry.teacher and entry.teacher.user == request.user:
        return redirect('teacher_visits')
    return redirect('visit_program_list')


@login_required
def visit_program_list(request):
    if not has_perm(request.user, 'visit_program', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().order_by('full_name')
    entries = VisitProgram.objects.select_related('teacher')
    if request.method == 'POST' and has_perm(request.user, 'visit_program', 'add'):
        teacher_id = request.POST.get('teacher_id')
        teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None
        if teacher:
            VisitProgram.objects.create(
                teacher=teacher,
                visit_date=request.POST.get('visit_date') or date.today(),
                lesson=request.POST.get('lesson', ''),
                notes=request.POST.get('notes', ''),
                created_by=request.user,
            )
            messages.success(request, 'تم إضافة السجل بنجاح')
            return redirect('visit_program_list')
        messages.error(request, 'الرجاء اختيار معلم')
    return render(request, 'school/visit_program_list.html', {
        'teachers': teachers,
        'entries': entries,
        'today': date.today(),
    })


@login_required
def visit_program_delete(request, entry_id):
    if not has_perm(request.user, 'visit_program', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    entry = get_object_or_404(VisitProgram, id=entry_id)
    entry.delete()
    messages.success(request, 'تم حذف السجل بنجاح')
    return redirect('visit_program_list')


@login_required
def visit_program_update(request, entry_id):
    if not has_perm(request.user, 'visit_program', 'add'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    entry = get_object_or_404(VisitProgram, id=entry_id)
    if request.method == 'POST':
        entry.notes = request.POST.get('notes', '')
        visit_date = request.POST.get('visit_date', '')
        lesson = request.POST.get('lesson', '')
        if visit_date:
            entry.visit_date = visit_date
        entry.lesson = lesson
        entry.save()
        messages.success(request, 'تم حفظ التعديل بنجاح')
    return redirect('visit_program_list')


@login_required
def visit_program_report(request):
    if not has_perm(request.user, 'visit_program', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    entries = VisitProgram.objects.select_related('teacher').order_by('-visit_date', 'teacher__full_name')
    info = SchoolInfo.objects.first()
    return render(request, 'school/visit_program_report.html', {
        'entries': entries,
        'info': info,
    })


@login_required
def visit_program_missing_notes_report(request):
    if not has_perm(request.user, 'visit_program', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    entries = VisitProgram.objects.filter(Q(notes__isnull=True) | Q(notes='')).filter(visit_date__gte=date.today()).select_related('teacher').order_by('visit_date', 'teacher__full_name')
    info = SchoolInfo.objects.first()
    return render(request, 'school/visit_program_report.html', {
        'entries': entries,
        'info': info,
        'missing_only': True,
    })


# ─── Absence (غياب الطلاب) ─────────────────────────────────────────────────────

@login_required
def absence_list(request):
    if not has_perm(request.user, 'absence', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    absence_date = request.GET.get('date') or date.today().isoformat()
    class_id = request.GET.get('class_id', '')
    selected_class = Class.objects.filter(id=class_id).first() if class_id else None
    if request.method == 'POST' and has_perm(request.user, 'absence', 'add'):
        absence_date = request.POST.get('absence_date') or date.today().isoformat()
        selected_class = Class.objects.filter(id=request.POST.get('class_id', '')).first()
        if selected_class:
            selected_ids = set()
            for sid in request.POST.getlist('student_id'):
                try:
                    selected_ids.add(int(sid))
                except ValueError:
                    continue
            StudentAbsence.objects.filter(
                student__student_class=selected_class, absence_date=absence_date,
            ).delete()
            students = Student.objects.filter(student_class=selected_class)
            to_create = [StudentAbsence(student=s, absence_date=absence_date, created_by=request.user)
                         for s in students if s.id in selected_ids]
            StudentAbsence.objects.bulk_create(to_create)
            for s in students:
                if s.id in selected_ids and s.user:
                    try:
                        Notification.objects.create(
                            user=s.user,
                            title='تنبيه بغياب الطالب',
                            message=f'الطالب {s.full_name} مسجَّل غائب بتاريخ {absence_date}. يرجى مراجعة المدرسة أو التواصل مع الإدارة.',
                            link='/leaves/',
                        )
                    except Exception:
                        pass
            log_action(request.user, 'تسجيل غياب طلاب', f'{len(to_create)} طالب - {absence_date} - الصف: {selected_class.name}')
            messages.success(request, f'تم حفظ غياب {len(to_create)} طالب بتاريخ {absence_date}')
        else:
            messages.error(request, 'الرجاء اختيار صف')
        return redirect(f'{request.path}?class_id={selected_class.id if selected_class else ""}&date={absence_date}')
    absent_ids = set()
    if selected_class:
        absent_ids = set(StudentAbsence.objects.filter(
            absence_date=absence_date, student__student_class=selected_class,
        ).values_list('student_id', flat=True))
    students = sort_students(Student.objects.filter(student_class=selected_class)) if selected_class else Student.objects.none()
    today = date.today()
    return render(request, 'school/absence_list.html', {
        'classes': classes,
        'selected_class': selected_class,
        'absence_date': absence_date,
        'absent_ids': absent_ids,
        'students': students,
        'today': today,
    })


@login_required
def absence_report(request):
    if not has_perm(request.user, 'absence', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    absence_date = request.GET.get('date') or date.today().isoformat()
    classes = Class.objects.all().order_by('name')
    rows = []
    for i, cls in enumerate(classes, start=1):
        absent = list(StudentAbsence.objects.filter(
            absence_date=absence_date, student__student_class=cls,
        ).select_related('student'))
        absent = sort_by_student_name(absent)
        class_students = Student.objects.filter(student_class=cls).count()
        rows.append({
            'num': i,
            'class_name': cls.name,
            'students_count': class_students,
            'present': class_students - len(absent),
            'count': len(absent),
            'absence_pct': round((len(absent) / class_students) * 100, 1) if class_students else 0,
            'names': ', '.join(a.student.full_name for a in absent) if absent else 'لا يوجد غياب',
        })
    total_students = Student.objects.count()
    total_absent = sum(r['count'] for r in rows)
    total_present = total_students - total_absent
    attendance_pct = round((total_present / total_students) * 100, 1) if total_students else 0
    absence_pct = round((total_absent / total_students) * 100, 1) if total_students else 0
    info = SchoolInfo.objects.first()
    return render(request, 'school/absence_report.html', {
        'rows': rows,
        'total_students': total_students,
        'total_absent': total_absent,
        'total_present': total_present,
        'attendance_pct': attendance_pct,
        'absence_pct': absence_pct,
        'absence_date': absence_date,
        'info': info,
        'today': date.today(),
    })


# ─── Teacher Daily Schedule (الجدول اليومي للمعلمين) ──────────────────────────

SCHEDULE_DAYS = ['الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
SCHEDULE_PERIODS = 7


@login_required
def schedule_edit(request):
    if not has_perm(request.user, 'schedule', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    is_teacher = request.user.profile.role == 'teacher'
    own_teacher = getattr(request.user, 'teacher_profile', None) if is_teacher else None
    if is_teacher:
        teacher = own_teacher
    else:
        teacher_id = request.GET.get('teacher_id', '') or request.POST.get('teacher_id', '')
        teacher = Teacher.objects.filter(id=teacher_id).first() if teacher_id and teacher_id.isdigit() else None
        teacher = teacher or Teacher.objects.first()
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'save' and has_perm(request.user, 'schedule', 'add'):
            if is_teacher:
                teacher = own_teacher
                if not teacher:
                    messages.error(request, 'لا يوجد حساب معلم مرتبط بحسابك')
                    return redirect('schedule_edit')
            if teacher:
                for day in SCHEDULE_DAYS:
                    for period in range(1, SCHEDULE_PERIODS + 1):
                        subject_id = request.POST.get(f'cell_{day}_{period}_subject', '')
                        class_id = request.POST.get(f'cell_{day}_{period}_class', '')
                        entry = TeacherScheduleEntry.objects.filter(teacher=teacher, day=day, period=period).first()
                        if subject_id or class_id:
                            TeacherScheduleEntry.objects.update_or_create(
                                teacher=teacher, day=day, period=period,
                                defaults={
                                    'subject_id': subject_id or None,
                                    'student_class_id': class_id or None,
                                    'updated_by': request.user,
                                },
                            )
                        elif entry:
                            entry.delete()
                messages.success(request, f'تم حفظ جدول {teacher.full_name}')
        if is_teacher:
            return redirect('schedule_edit')
        return redirect(f'{request.path}?teacher_id={teacher.id}')
    classes = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')
    teachers = [] if is_teacher else Teacher.objects.all().order_by('full_name')
    entries = list(TeacherScheduleEntry.objects.filter(teacher=teacher).select_related('subject', 'student_class')) if teacher else []
    return render(request, 'school/schedule_edit.html', {
        'teachers': teachers,
        'teacher': teacher,
        'is_teacher': is_teacher,
        'days': SCHEDULE_DAYS,
        'period_range': range(1, SCHEDULE_PERIODS + 1),
        'subjects': subjects,
        'classes': classes,
        'entries': entries,
    })


def _schedule_info():
    return SchoolInfo.objects.first()


@login_required
def schedule_print_cards(request):
    if not has_perm(request.user, 'schedule', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    is_teacher = request.user.profile.role == 'teacher'
    own_teacher = getattr(request.user, 'teacher_profile', None) if is_teacher else None
    teachers = Teacher.objects.all().order_by('full_name')
    entries = list(TeacherScheduleEntry.objects.all().select_related('teacher', 'subject', 'student_class'))
    if is_teacher and own_teacher:
        teachers = [own_teacher]
        entries = [e for e in entries if e.teacher_id == own_teacher.id]
    return render(request, 'school/schedule_print_cards.html', {
        'teachers': teachers,
        'entries': entries,
        'days': SCHEDULE_DAYS,
        'period_range': range(1, SCHEDULE_PERIODS + 1),
        'info': _schedule_info(),
        'today': date.today(),
    })


@login_required
def schedule_print_all(request):
    if not has_perm(request.user, 'schedule', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    is_teacher = request.user.profile.role == 'teacher'
    own_teacher = getattr(request.user, 'teacher_profile', None) if is_teacher else None
    teachers = Teacher.objects.all().order_by('full_name')
    entries = list(TeacherScheduleEntry.objects.all().select_related('teacher', 'subject', 'student_class'))
    if is_teacher and own_teacher:
        teachers = [own_teacher]
        entries = [e for e in entries if e.teacher_id == own_teacher.id]
    return render(request, 'school/schedule_print_all.html', {
        'teachers': teachers,
        'entries': entries,
        'days': SCHEDULE_DAYS,
        'period_range': range(1, SCHEDULE_PERIODS + 1),
        'info': _schedule_info(),
        'today': date.today(),
    })


@login_required
def schedule_print_classes(request):
    if not has_perm(request.user, 'schedule', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    is_teacher = request.user.profile.role == 'teacher'
    own_teacher = getattr(request.user, 'teacher_profile', None) if is_teacher else None
    classes = Class.objects.all().order_by('name')
    entries = list(TeacherScheduleEntry.objects.filter(student_class__isnull=False).select_related('teacher', 'subject', 'student_class'))
    if is_teacher and own_teacher:
        entries = [e for e in entries if e.teacher_id == own_teacher.id]
    return render(request, 'school/schedule_print_classes.html', {
        'classes': classes,
        'entries': entries,
        'days': SCHEDULE_DAYS,
        'period_range': range(1, SCHEDULE_PERIODS + 1),
        'info': _schedule_info(),
        'today': date.today(),
    })


@login_required
def schedule_print_admin(request):
    if not has_perm(request.user, 'schedule', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    is_teacher = request.user.profile.role == 'teacher'
    own_teacher = getattr(request.user, 'teacher_profile', None) if is_teacher else None
    classes = list(Class.objects.all().order_by('name'))
    entries = list(TeacherScheduleEntry.objects.filter(student_class__isnull=False).select_related('teacher', 'subject', 'student_class'))
    if is_teacher and own_teacher:
        entries = [e for e in entries if e.teacher_id == own_teacher.id]
    return render(request, 'school/schedule_print_admin.html', {
        'classes': classes,
        'entries': entries,
        'days': SCHEDULE_DAYS,
        'period_range': range(1, SCHEDULE_PERIODS + 1),
        'info': _schedule_info(),
        'today': date.today(),
    })


# ─── Health & Social Survey (المسح الصحي والاجتماعي) ─────────────────────────

@login_required
def survey_form(request):
    student = getattr(request.user, 'student_profile', None)
    if not student:
        messages.error(request, 'هذه الصفحة خاصة بحسابات الطلاب')
        return redirect('dashboard')
    if not has_perm(request.user, 'survey', 'add'):
        messages.error(request, 'المسح مغلق حالياً')
        return redirect('dashboard')
    survey = StudentSurvey.objects.filter(student=student).first()
    if request.method == 'POST':
        if not survey:
            survey = StudentSurvey(student=student)
        form = StudentSurveyForm(request.POST, instance=survey)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم حفظ بيانات المسح بنجاح')
            return redirect('survey_form')
    else:
        form = StudentSurveyForm(instance=survey)
    return render(request, 'school/survey_form.html', {
        'student': student,
        'form': form,
        'has_survey': bool(survey),
        'submitted_at': survey.submitted_at if survey else None,
    })


@login_required
def survey_report(request):
    if not has_perm(request.user, 'survey', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    surveys = StudentSurvey.objects.select_related('student__student_class')
    total_students = Student.objects.count()
    submitted_ids = set(surveys.values_list('student_id', flat=True))
    not_submitted = sort_students_class_first(Student.objects.exclude(id__in=submitted_ids).select_related('student_class'))
    classes = Class.objects.all().order_by('name')
    stats = []
    for cls in classes:
        total = Student.objects.filter(student_class=cls).count()
        done = surveys.filter(student__student_class=cls).count()
        stats.append({'class': cls, 'total': total, 'done': done, 'missing': total - done})
    surveys_sorted = sorted(list(surveys), key=lambda s: (
        arabic_sort_key(s.student.student_class.name) if s.student.student_class else (0,),
        arabic_sort_key(s.student.full_name),
    ))
    return render(request, 'school/survey_report.html', {
        'surveys': surveys_sorted,
        'total_students': total_students,
        'submitted_count': len(submitted_ids),
        'not_submitted_count': len(not_submitted),
        'not_submitted': not_submitted,
        'stats': stats,
        'today': date.today(),
    })


def survey_condition_labels(s):
    out = []
    if s.chronic_disease:
        out.append('مرض مزمن' + (f' ({s.chronic_disease_details})' if s.chronic_disease_details else ''))
    if s.condition_asthma:
        out.append('الربو')
    if s.condition_diabetes:
        out.append('السكري')
    if s.condition_epilepsy:
        out.append('الصرع')
    if s.condition_heart:
        out.append('مشاكل قلب')
    if s.condition_hearing:
        out.append('ضعف سمع')
    if s.condition_vision:
        out.append('ضعف بصر')
    if s.allergy_drugs:
        out.append('حساسية أدوية')
    if s.allergy_food:
        out.append('حساسية أطعمة')
    if s.allergy_dust:
        out.append('حساسية غبار')
    if s.regular_medication:
        out.append('أدوية منتظمة')
    if s.special_care:
        out.append('رعاية خاصة')
    if s.family_special_conditions:
        out.append('ظروف أسرية')
    return '، '.join(out) or '-'


def build_survey_stats_data():
    """يحسب كل إحصاءات المسح ويُرجع قاموساً لاستخدامه في الصفحة أو تقارير الطباعة."""
    surveys = StudentSurvey.objects.select_related('student__student_class')
    submitted = list(surveys)
    submitted.sort(key=lambda s: (
        arabic_sort_key(s.student.student_class.name) if s.student.student_class else (0,),
        arabic_sort_key(s.student.full_name),
    ))
    n = max(len(submitted), 1)

    def cnt(pred):
        return len([s for s in submitted if pred(s)])

    def pct(c):
        return round(c * 100.0 / n, 1)

    # ── مؤشرات صحية ──
    health_rows = []
    indicators = [
        ('مرض مزمن (أي نوع)', lambda s: s.chronic_disease, 'danger'),
        ('أدوية منتظمة', lambda s: s.regular_medication, 'warning'),
        ('حساسية (أي نوع)', lambda s: s.has_allergy, 'warning'),
        ('الربو', lambda s: s.condition_asthma, 'danger'),
        ('السكري', lambda s: s.condition_diabetes, 'danger'),
        ('الصرع', lambda s: s.condition_epilepsy, 'danger'),
        ('مشاكل في القلب', lambda s: s.condition_heart, 'danger'),
        ('ضعف في السمع', lambda s: s.condition_hearing, 'secondary'),
        ('ضعف في البصر', lambda s: s.condition_vision, 'secondary'),
        ('يحتاج نظارات طبية', lambda s: s.needs_glasses, 'info'),
        ('يحتاج رعاية صحية خاصة أثناء الدوام', lambda s: s.special_care, 'danger'),
        ('لديه تعليمات طارئة مسجلة', lambda s: bool((s.emergency_instructions or '').strip()), 'info'),
    ]
    for label, fn, color in indicators:
        c = cnt(fn)
        health_rows.append({'label': label, 'count': c, 'pct': pct(c), 'color': color, 'students': [s for s in submitted if fn(s)]})

    # ── مؤشرات اجتماعية ──
    social_rows = [
        {'label': 'صعوبات تؤثر على الدراسة', 'count': 0, 'pct': 0, 'color': 'warning', 'students': []},
    ]
    social_rows[0]['count'] = cnt(lambda s: s.study_difficulties)
    social_rows[0]['pct'] = pct(social_rows[0]['count'])
    social_rows[0]['students'] = [s for s in submitted if s.study_difficulties]

    supports = []
    for label, fn in [
        ('يحتاج دعم أكاديمي', lambda s: s.support_academic),
        ('يحتاج دعم نفسي', lambda s: s.support_psychological),
        ('يحتاج دعم اجتماعي', lambda s: s.support_social),
        ('لا يحتاج دعم', lambda s: s.support_none),
    ]:
        c = cnt(fn)
        supports.append({'label': label, 'count': c, 'pct': pct(c), 'students': [s for s in submitted if fn(s)]})

    device_rows = []
    for label, fn in [
        ('هاتف ذكي', lambda s: s.has_smartphone),
        ('جهاز حاسوب', lambda s: s.has_computer),
        ('اتصال بالإنترنت', lambda s: s.has_internet),
        ('لا يتوفر أي من الأجهزة', lambda s: s.has_no_device),
    ]:
        c = cnt(fn)
        device_rows.append({'label': label, 'count': c, 'pct': pct(c)})

    living_rows = []
    for val, label in [('parents', 'الأب والأم'), ('father', 'الأب فقط'), ('mother', 'الأم فقط'), ('relative', 'أحد الأقارب'), ('other', 'أخرى')]:
        c = len([s for s in submitted if s.lives_with == val])
        living_rows.append({'value': val, 'label': label, 'count': c, 'pct': pct(c)})

    # ── طلاب بحالة تستدعي العناية ──
    def is_critical(s):
        return (s.chronic_disease or s.condition_asthma or s.condition_diabetes or s.condition_epilepsy
                or s.condition_heart or s.condition_hearing or s.condition_vision
                or s.has_allergy or s.regular_medication or s.special_care or s.family_special_conditions)

    critical = [{'survey': s, 'labels': survey_condition_labels(s)} for s in submitted if is_critical(s)]
    critical.sort(key=lambda d: (d['survey'].student.student_class.name if d['survey'].student.student_class else ''))

    # ── ملخص حسب الصف ──
    class_stats = []
    for cls in Class.objects.all().order_by('name'):
        cs = [s for s in submitted if s.student.student_class_id == cls.id]
        class_stats.append({
            'class': cls,
            'done': len(cs),
            'chronic': len([s for s in cs if s.chronic_disease]),
            'meds': len([s for s in cs if s.regular_medication]),
            'allergy': len([s for s in cs if s.has_allergy]),
            'conditions': len([s for s in cs if is_critical(s)]),
            'study_diff': len([s for s in cs if s.study_difficulties]),
        })

    return {
        'submitted_count': n,
        'health_rows': health_rows,
        'social_rows': social_rows,
        'supports': supports,
        'device_rows': device_rows,
        'living_rows': living_rows,
        'critical': critical,
        'class_stats': class_stats,
        'submitted': submitted,
        'today': date.today(),
    }


@login_required
def survey_stats(request):
    if not has_perm(request.user, 'survey', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    data = build_survey_stats_data()
    return render(request, 'school/survey_stats.html', data)


@login_required
def survey_stats_print(request, mode):
    """تقرير قابل للطباعة: students = قائمة بكل الطلاب، groups = تقسيمات حسب الحالة."""
    if not has_perm(request.user, 'survey', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    data = build_survey_stats_data()
    info = SchoolInfo.objects.first()

    if mode == 'students':
        rows = []
        for s in data['submitted']:
            rows.append({
                'survey': s,
                'name': s.student.full_name,
                'cls': s.student.student_class.name if s.student.student_class else '-',
                'labels': survey_condition_labels(s),
            })
        return render(request, 'school/survey_stats_print.html', {
            'mode': 'students',
            'rows': rows,
            'info': info,
            'submitted_count': data['submitted_count'],
            'today': data['today'],
        })

    groups = []
    for r in data['health_rows'] + data['social_rows'] + data['supports']:
        if r['students']:
            groups.append({'title': r['label'], 'students': r['students']})
    return render(request, 'school/survey_stats_print.html', {
        'mode': 'groups',
        'groups': groups,
        'info': info,
        'submitted_count': data['submitted_count'],
        'today': data['today'],
    })


@login_required
def survey_detail(request, student_id):
    if not has_perm(request.user, 'survey', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    student = get_object_or_404(Student, id=student_id)
    survey = StudentSurvey.objects.filter(student=student).first()
    return render(request, 'school/survey_detail.html', {
        'student': student,
        'survey': survey,
        'today': date.today(),
    })


# ─── Guardians / Nominations / Certificates ───────────────────────────────────

def current_teaching_year():
    today = date.today()
    y = today.year
    return f'{y}/{y + 1}' if today.month >= 9 else f'{y - 1}/{y}'


def issue_certificates(nominations, user):
    info = SchoolInfo.objects.first()
    year = current_teaching_year()
    count = 0
    for n in nominations:
        if Certificate.objects.filter(nomination=n).exists():
            continue
        Certificate.objects.create(
            nomination=n,
            student_name=n.student.full_name,
            class_name=n.student_class.name,
            guardian_name=n.student_class.guardian.full_name if n.student_class.guardian else '',
            principal_name=info.principal_name if info else '',
            cert_year=year,
            issued_by=user,
        )
        count += 1
    return count


@login_required
def guardians_report(request):
    if not has_perm(request.user, 'guardians', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    info = SchoolInfo.objects.first()
    class_rows = []
    for i, cls in enumerate(Class.objects.all().order_by('name'), start=1):
        class_rows.append({
            'num': i,
            'name': cls.name,
            'students_count': Student.objects.filter(student_class=cls).count(),
            'guardian': cls.guardian,
        })
    unassigned = Teacher.objects.filter(guardian_class=None).order_by('full_name')
    return render(request, 'school/guardians_report.html', {
        'class_rows': class_rows,
        'unassigned': unassigned,
        'info': info,
        'today': date.today(),
    })


@login_required
def guardian_assign(request):
    if not has_perm(request.user, 'guardians', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().select_related('guardian_class').order_by('full_name')
    classes = Class.objects.all().order_by('name')
    if request.method == 'POST' and has_perm(request.user, 'guardians', 'add'):
        for t in teachers:
            cid = request.POST.get(f'guardian_{t.id}', '')
            current = Class.objects.filter(guardian=t).first()
            if cid:
                cls = get_object_or_404(Class, id=cid)
                if current and current.id != cls.id:
                    current.guardian = None
                    current.save()
                cls.guardian = t
                cls.save()
            elif current:
                current.guardian = None
                current.save()
        messages.success(request, 'تم حفظ مربي الصفوف بنجاح')
        return redirect('guardian_assign')
    return render(request, 'school/guardian_assign.html', {
        'teachers': teachers,
        'classes': classes,
    })


@login_required
def guardian_students(request):
    is_admin = request.user.profile.role == 'admin'
    if not is_admin and request.user.profile.role not in ('teacher', 'vice_principal', 'secretary'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teacher = getattr(request.user, 'teacher_profile', None)
    if is_admin:
        classes = Class.objects.all().order_by('name')
        class_id = request.POST.get('class_id') or request.GET.get('class_id', '')
        guardian_class = Class.objects.filter(id=class_id).first() if class_id else classes.first()
        if not guardian_class:
            messages.error(request, 'لا توجد صفوف دراسية بعد')
            return redirect('dashboard')
    else:
        classes = []
        guardian_class = getattr(teacher, 'guardian_class', None) if teacher else None
        if not guardian_class:
            messages.error(request, 'لم يتم تخصيص صف لك كمربي صف بعد')
            return redirect('dashboard')
    students = sort_students(Student.objects.filter(student_class=guardian_class))
    nominated_ids = set(Nomination.objects.filter(student_class=guardian_class).values_list('student_id', flat=True))
    can_nominate = has_perm(request.user, 'nominations', 'add') or is_admin
    if request.method == 'POST':
        if not can_nominate:
            messages.error(request, 'ليس لديك صلاحية الترشيح')
        else:
            selected = set()
            for sid in request.POST.getlist('student_id'):
                try:
                    selected.add(int(sid))
                except ValueError:
                    continue
            Nomination.objects.filter(student_class=guardian_class).delete()
            for sid in selected:
                st = Student.objects.filter(id=sid, student_class=guardian_class).first()
                if st:
                    Nomination.objects.create(student=st, student_class=guardian_class, nominated_by=request.user)
            messages.success(request, f'تم حفظ ترشيحات {len(selected)} طالب كمتفوقين')
        if is_admin:
            return redirect(f'/guardians/students/?class_id={guardian_class.id}')
        return redirect('guardian_students')
    return render(request, 'school/guardian_students.html', {
        'students': students,
        'guardian_class': guardian_class,
        'classes': classes,
        'is_admin': is_admin,
        'nominated_ids': nominated_ids,
        'today': date.today(),
    })


@login_required
def certificates_manage(request):
    if not has_perm(request.user, 'certificates', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    classes = Class.objects.all().order_by('name')
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'add' and has_perm(request.user, 'certificates', 'add'):
            student_id = request.POST.get('student_id')
            student = get_object_or_404(Student, id=student_id) if student_id else None
            if student and student.student_class:
                if Nomination.objects.filter(student=student).exists():
                    messages.error(request, 'هذا الطالب مرشح مسبقاً')
                else:
                    Nomination.objects.create(student=student, student_class=student.student_class, nominated_by=request.user)
                    messages.success(request, 'تمت إضافة الطالب يدوياً')
            else:
                messages.error(request, 'الرجاء اختيار طالب له صف')
        elif action == 'delete' and has_perm(request.user, 'certificates', 'delete'):
            Nomination.objects.filter(id=request.POST.get('nomination_id', '')).delete()
            messages.success(request, 'تم حذف الترشيح')
        elif action == 'issue' and has_perm(request.user, 'certificates', 'add'):
            nominations = Nomination.objects.select_related('student', 'student_class__guardian')
            class_id = request.POST.get('class_id', '')
            if class_id:
                nominations = nominations.filter(student_class_id=class_id)
            count = issue_certificates(nominations, request.user)
            messages.success(request, f'تم إصدار {count} شهادة')
        return redirect('certificates_manage')
    class_id = request.GET.get('class_id', '')
    nominations = Nomination.objects.select_related('student', 'student_class__guardian', 'nominated_by')
    if class_id:
        nominations = nominations.filter(student_class_id=class_id)
    students = sort_students_class_first(Student.objects.all().select_related('student_class'))
    nominated_classes = Class.objects.filter(nominations__isnull=False).distinct().order_by('name')
    classes_without = Class.objects.exclude(id__in=nominated_classes.values('id')).order_by('name')
    return render(request, 'school/certificates_manage.html', {
        'nominations': nominations,
        'classes': classes,
        'nominated_classes': nominated_classes,
        'classes_without': classes_without,
        'students': students,
        'selected_class_id': class_id,
        'year': current_teaching_year(),
    })


@login_required
def certificate_print(request, nomination_id):
    if not has_perm(request.user, 'certificates', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    nomination = get_object_or_404(Nomination, id=nomination_id)
    issue_certificates(Nomination.objects.filter(id=nomination.id), request.user)
    cert = Certificate.objects.get(nomination=nomination)
    info = SchoolInfo.objects.first()
    bg = request.GET.get('bg', '1')
    return render(request, 'school/certificate_print.html', {
        'cert': cert,
        'info': info,
        'bg': bg,
    })


@login_required
def certificates_print_all(request):
    if not has_perm(request.user, 'certificates', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    nominations = Nomination.objects.select_related('student', 'student_class__guardian')
    class_id = request.GET.get('class_id', '')
    if class_id:
        nominations = nominations.filter(student_class_id=class_id)
    issue_certificates(nominations, request.user)
    nominations = sorted(nominations, key=lambda n: (arabic_sort_key(n.student_class.name), arabic_sort_key(n.student.full_name)))
    certs = [Certificate.objects.get(nomination=n) for n in nominations]
    info = SchoolInfo.objects.first()
    bg = request.GET.get('bg', '1')
    return render(request, 'school/certificates_print_all.html', {
        'certs': certs,
        'info': info,
        'bg': bg,
    })


@login_required
def certificates_export(request):
    if not has_perm(request.user, 'certificates', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    import openpyxl
    nominations = Nomination.objects.select_related('student', 'student_class__guardian')
    class_id = request.GET.get('class_id', '')
    if class_id:
        nominations = nominations.filter(student_class_id=class_id)
    info = SchoolInfo.objects.first()
    principal = info.principal_name if info else ''
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المتفوقون'
    headers = ['اسم الطالب', 'الصف', 'اسم مدير المدرسة', 'اسم مربي الصف']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    for n in sorted(nominations, key=lambda n: arabic_sort_key(n.student.full_name)):
        ws.append([
            n.student.full_name,
            n.student_class.name,
            principal,
            n.student_class.guardian.full_name if n.student_class.guardian else '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=المتفوقون.xlsx'
    wb.save(response)
    return response


# ─── PWA & Web Push ───────────────────────────────────────────────────────────

import json as json_lib
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse


@login_required
def push_subscribe(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST فقط'}, status=405)
    try:
        data = json_lib.loads(request.body.decode('utf-8'))
        endpoint = data.get('endpoint', '')
        p256dh = data.get('keys', {}).get('p256dh', '')
        auth = data.get('keys', {}).get('auth', '')
        if not (endpoint and p256dh and auth):
            return JsonResponse({'ok': False, 'error': 'بيانات ناقصة'}, status=400)
        PushSubscription.objects.update_or_create(
            endpoint=endpoint,
            defaults={'user': request.user, 'p256dh': p256dh, 'auth': auth},
        )
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
def push_unsubscribe(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST فقط'}, status=405)
    try:
        data = json_lib.loads(request.body.decode('utf-8'))
        PushSubscription.objects.filter(user=request.user, endpoint=data.get('endpoint', '')).delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
def push_test(request):
    sent = send_push(request.user, 'اختبار الإشعارات', 'إشعارات النظام تعمل بنجاح', '/dashboard/')
    if sent:
        messages.success(request, 'تم إرسال إشعار تجريبي لجهازك')
    else:
        messages.error(request, 'لا يوجد اشتراك مفعّل لهذا الحساب، أو لم تُضبط مفاتيح VAPID')
    return redirect('dashboard')


def service_worker(request):
    response = render(request, 'school/sw.js', content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    response['Cache-Control'] = 'no-cache'
    return response


def manifest_view(request):
    info = SchoolInfo.objects.first()
    return JsonResponse({
        'name': info.name_ar if info else 'النظام المدرسي',
        'short_name': 'النظام المدرسي',
        'start_url': '/dashboard/',
        'display': 'standalone',
        'dir': 'rtl',
        'lang': 'ar',
        'background_color': '#0d1b2a',
        'theme_color': '#0d1b2a',
        'icons': [
            {'src': '/static/pwa/icon-192.png', 'sizes': '192x192', 'type': 'image/png', 'purpose': 'any maskable'},
            {'src': '/static/pwa/icon-512.png', 'sizes': '512x512', 'type': 'image/png', 'purpose': 'any maskable'},
        ],
    })


# ─── Notifications ─────────────────────────────────────────────────────────────

@login_required
def notification_list(request):
    notifications = Notification.objects.filter(user=request.user).exclude(link__startswith='/messages/')
    notifications.filter(is_read=False).update(is_read=True)
    return render(request, 'school/notification_list.html', {
        'notifications': notifications,
    })


@login_required
def notification_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    if notification.link:
        link = notification.link
        if link.startswith('/messages/'):
            match = re.match(r'^/messages/(\d+)/', link)
            if not match or not Message.objects.filter(id=int(match.group(1))).exists():
                notification.delete()
                messages.error(request, 'هذه الرسالة لم تعد موجودة')
                return redirect('notification_list')
        return redirect(link)
    return redirect('notification_list')


# ─── Inspection Visits (Principal) ─────────────────────────────────────────────

@login_required
def inspection_visit_list(request):
    if not has_perm(request.user, 'inspection_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    teachers = Teacher.objects.all().order_by('full_name')
    selected_teacher = None
    visits = InspectionVisit.objects.none()
    teacher_id = request.GET.get('teacher_id', '')
    if teacher_id:
        selected_teacher = get_object_or_404(Teacher, id=teacher_id)
        visits = InspectionVisit.objects.filter(teacher=selected_teacher).order_by('-visit_date')
    if request.method == 'POST' and has_perm(request.user, 'inspection_visits', 'add'):
        teacher_id = request.POST.get('teacher_id')
        selected_teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None
        if selected_teacher:
            visit = InspectionVisit.objects.create(
                teacher=selected_teacher,
                visit_date=request.POST.get('visit_date') or date.today(),
                visit_number=request.POST.get('visit_number', ''),
                subject_area=request.POST.get('subject_area', ''),
                lesson_topic=request.POST.get('lesson_topic', ''),
                class_name=request.POST.get('class_name', ''),
                section=request.POST.get('section', ''),
                content_teaching=request.POST.get('content_teaching', ''),
                teaching_strategies=request.POST.get('teaching_strategies', ''),
                evaluation_strategies=request.POST.get('evaluation_strategies', ''),
                other_matters=request.POST.get('other_matters', ''),
                plans_followup=request.POST.get('plans_followup', ''),
                attendance_followup=request.POST.get('attendance_followup', ''),
                committees_followup=request.POST.get('committees_followup', ''),
                violence_policy=request.POST.get('violence_policy', ''),
                recommendations=request.POST.get('recommendations', ''),
                principal_sign_date=request.POST.get('principal_sign_date') or date.today(),
                teacher_receipt_date=request.POST.get('teacher_receipt_date') or date.today(),
                created_by=request.user,
            )
            if selected_teacher.user:
                Notification.objects.create(
                    user=selected_teacher.user,
                    title='زيارة إشرافية جديدة',
                    message=f'تم تسجيل زيارة إشرافية بتاريخ {visit.visit_date}',
                    link=f'/inspection-visits/{visit.id}/report/',
                )
                send_push(selected_teacher.user, 'زيارة إشرافية جديدة', f'تم تسجيل زيارة إشرافية بتاريخ {visit.visit_date}', f'/inspection-visits/{visit.id}/report/')
            messages.success(request, 'تم إضافة الزيارة الإشرافية بنجاح')
            return redirect(f'{request.path}?teacher_id={teacher_id}')
        messages.error(request, 'الرجاء اختيار معلم')
    return render(request, 'school/inspection_visit_list.html', {
        'teachers': teachers,
        'selected_teacher': selected_teacher,
        'visits': visits,
        'today': date.today(),
    })


@login_required
def inspection_visit_report(request, visit_id):
    if not has_perm(request.user, 'inspection_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(InspectionVisit, id=visit_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/inspection_visit_report.html', {
        'visit': visit,
        'info': info,
    })


@login_required
def inspection_visits_all_report(request):
    if not has_perm(request.user, 'inspection_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visits = InspectionVisit.objects.all().select_related('teacher').order_by('-visit_date')
    info = SchoolInfo.objects.first()
    return render(request, 'school/inspection_visits_all_report.html', {
        'visits': visits,
        'info': info,
    })


# ─── Backup / Restore ─────────────────────────────────────────────────────────

@login_required
def backup_data(request):
    """نسخ احتياطي لكل البيانات (JSON) أو استعادتها — للمدير فقط."""
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')

    import tempfile
    from django.core.management import call_command

    EXCLUDED = ['contenttypes', 'sessions', 'admin']

    if request.method == 'POST':
        restore_file = request.FILES.get('backup_file')
        if not restore_file or not restore_file.name.endswith('.json'):
            messages.error(request, 'يرجى اختيار ملف JSON صالح')
            return redirect('backup_data')
        tmp = tempfile.NamedTemporaryFile(suffix='.json', delete=False, dir='/tmp' if os.getenv('VERCEL') else None)
        try:
            tmp.write(restore_file.read())
            tmp.close()
            call_command('loaddata', tmp.name, verbosity=0)
            messages.success(request, 'تمت استعادة النسخة الاحتياطية بنجاح')
        except Exception as e:
            messages.error(request, f'فشلت الاستعادة: {e}')
        finally:
            if os.path.exists(tmp.name):
                os.remove(tmp.name)
        return redirect('backup_data')

    if request.GET.get('download') != '1':
        return render(request, 'school/backup_data.html')

    out = io.StringIO()
    call_command('dumpdata', 'school', 'auth', exclude=EXCLUDED, stdout=out, verbosity=0)
    filename = f'backup_{date.today().strftime("%Y%m%d")}.json'
    response = HttpResponse(out.getvalue(), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# ─── Reset / Flush Data ───────────────────────────────────────────────────────

CLEARABLE_TABLES = [
    ('leaves', 'أذونات المغادرة', StudentLeave, ['student']),
    ('notes', 'ملاحظات الطلاب', Note, ['student']),
    ('lateness', 'تأخيرات الطلاب', StudentLateness, ['student']),
    ('absence', 'غياب الطلاب', StudentAbsence, ['student']),
    ('levels', 'مستويات الطلاب', StudentLevel, ['student']),
    ('survey', 'المسوح الصحية والاجتماعية', StudentSurvey, ['student']),
    ('visits', 'زيارات المشرفين', SupervisorVisit, ['teacher']),
    ('inspection_visits', 'الزيارات الإشرافية', InspectionVisit, ['teacher']),
    ('visit_program', 'برنامج الزيارات', VisitProgram, ['teacher']),
    ('meetings', 'اجتماعات المعلمين', Meeting, ['teacher']),
    ('exams', 'تحليل الامتحانات', ExamAnalysis, []),
    ('messages', 'الرسائل', Message, []),
    ('announcements', 'الإعلانات', Announcement, []),
    ('agenda', 'الأجندة', Agenda, []),
    ('schedule', 'الجدول اليومي للمعلمين', TeacherScheduleEntry, []),
    ('notifications', 'الإشعارات', Notification, []),
    ('certificates', 'شهادات التقدير', Certificate, []),
    ('nominations', 'ترشيحات المتفوقين', Nomination, []),
    ('teacher_notes', 'ملاحظات المعلمين', TeacherNote, []),
    ('lesson_links', 'روابط الدروس', LessonLink, []),
    ('whatsapp_groups', 'روابط واتساب الصفوف', WhatsAppGroup, []),
    ('incoming', 'سجل الوارد', IncomingLetter, ['created_by']),
    ('outgoing', 'سجل الصادر', OutgoingLetter, ['created_by']),
    ('teacher_followups', 'متابعة المعلمين', TeacherFollowup, ['teacher']),
    ('reciprocal_visits', 'الزيارات التبادلية', ReciprocalVisit, ['teacher']),
    ('no_objections', 'لا مانع', NoObjection, ['created_by']),
    ('warnings', 'إنذارات الطلاب', StudentWarning, ['student']),
    ('summons', 'استدعاءات أولياء الأمور', GuardianSummons, ['student']),
]

DEPENDENT_MODELS = {
    'students': [Note, StudentLeave, StudentLateness, StudentAbsence, StudentLevel, StudentSurvey, LoginCounter, StudentWarning, GuardianSummons],
    'teachers': [TeacherNote, Meeting, SupervisorVisit, InspectionVisit, VisitProgram, TeacherScheduleEntry, TeacherFollowup, ReciprocalVisit],
    'classes': [Student, TeacherScheduleEntry, WhatsAppGroup],
    'subjects': [TeacherScheduleEntry],
}


@login_required
def reset_data(request):
    if request.user.profile.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        action = request.POST.get('action', '')
        confirm = request.POST.get('confirm', '')
        if confirm != 'YES':
            messages.error(request, 'يرجى كتابة YES للتأكيد')
            return redirect('reset_data')
        if action == 'flush_one':
            key = request.POST.get('key', '')
            for k, label, model, deps in CLEARABLE_TABLES:
                if k == key:
                    log_action(request.user, 'تفريغ بيانات', label)
                    model.objects.all().delete()
                    messages.success(request, f'تم تفريغ: {label}')
                    return redirect('reset_data')
            messages.error(request, 'قسم غير معروف')
            return redirect('reset_data')
        if action == 'flush_related':
            key = request.POST.get('key', '')
            if key == 'students':
                log_action(request.user, 'تفريغ الطلاب', 'حذف جميع الطلاب وجميع بياناتهم المرتبطة')
                for m in DEPENDENT_MODELS['students']:
                    m.objects.all().delete()
                Student.objects.all().delete()
                User.objects.filter(profile__role='student').delete()
                messages.success(request, 'تم تفريغ الطلاب وجميع بياناتهم المرتبطة')
            elif key == 'teachers':
                log_action(request.user, 'تفريغ المعلمين', 'حذف جميع المعلمين وجميع بياناتهم المرتبطة')
                for m in DEPENDENT_MODELS['teachers']:
                    m.objects.all().delete()
                Teacher.objects.all().delete()
                User.objects.filter(profile__role='teacher').delete()
                messages.success(request, 'تم تفريغ المعلمين وجميع بياناتهم المرتبطة')
            elif key == 'classes':
                log_action(request.user, 'تفريغ الصفوف', 'حذف جميع الصفوف وجميع بياناتهم المرتبطة')
                for m in DEPENDENT_MODELS['classes']:
                    m.objects.all().delete()
                Class.objects.all().delete()
                messages.success(request, 'تم تفريغ الصفوف وجميع بياناتهم المرتبطة')
            elif key == 'subjects':
                log_action(request.user, 'تفريغ المواد', 'حذف جميع المواد وجميع بياناتهم المرتبطة')
                for m in DEPENDENT_MODELS['subjects']:
                    m.objects.all().delete()
                Subject.objects.all().delete()
                messages.success(request, 'تم تفريغ المواد وجميع بياناتهم المرتبطة')
            return redirect('reset_data')
    counts = [{'key': k, 'label': label, 'count': model.objects.count()} for k, label, model, deps in CLEARABLE_TABLES]
    return render(request, 'school/reset_data.html', {
        'counts': counts,
        'full_groups': [('students', 'الطلاب'), ('teachers', 'المعلمون'), ('classes', 'الصفوف'), ('subjects', 'المواد')],
    })


# ─── WhatsApp Group Links (روابط واتساب الصفوف) ──────────────────────────────

@login_required
def whatsapp_groups(request):
    if not has_perm(request.user, 'settings', 'whatsapp'):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('dashboard')
    if request.method == 'POST':
        class_id = request.POST.get('class_id', '')
        link = request.POST.get('link', '').strip()
        if not class_id or not link:
            messages.error(request, 'اختر الصف وأدخل رابط واتساب')
        else:
            cls = get_object_or_404(Class, id=class_id)
            WhatsAppGroup.objects.create(student_class=cls, link=link)
            messages.success(request, 'تم إضافة رابط واتساب بنجاح')
        return redirect('whatsapp_groups')
    groups = WhatsAppGroup.objects.select_related('student_class').all()
    classes = Class.objects.all().order_by('name')
    return render(request, 'school/whatsapp_groups.html', {
        'groups': groups,
        'classes': classes,
    })


@login_required
def whatsapp_group_delete(request, group_id):
    if not has_perm(request.user, 'settings', 'whatsapp'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    group = get_object_or_404(WhatsAppGroup, id=group_id)
    group.delete()
    messages.success(request, 'تم حذف رابط واتساب')
    return redirect('whatsapp_groups')


# ─── Secretary: Incoming (سجل الوارد) ────────────────────────────────────────

def current_academic_year():
    today = date.today()
    if today.month >= 8:
        return f'{today.year}/{today.year + 1}'
    return f'{today.year - 1}/{today.year}'


@login_required
def incoming_list(request):
    if not has_perm(request.user, 'incoming', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    year = request.GET.get('year', '').strip()
    if not year:
        year = current_academic_year()
    if request.method == 'POST' and has_perm(request.user, 'incoming', 'add'):
        number = request.POST.get('number', '').strip()
        if not number:
            messages.error(request, 'رقم الوارد مطلوب')
        else:
            IncomingLetter.objects.create(
                academic_year=request.POST.get('academic_year', '').strip() or year,
                number=number,
                date=request.POST.get('date', '') or date.today(),
                letter_type=request.POST.get('letter_type', '').strip(),
                source_entity=request.POST.get('source_entity', '').strip(),
                attachments=request.POST.get('attachments', '').strip(),
                subject=request.POST.get('subject', '').strip(),
                file_number=request.POST.get('file_number', '').strip(),
                created_by=request.user,
            )
            messages.success(request, 'تم إضافة الوارد بنجاح')
        return redirect(f'{reverse("incoming_list")}?year={year}')
    letters = IncomingLetter.objects.filter(academic_year=year).order_by('date', 'number')
    years = list(IncomingLetter.objects.values_list('academic_year', flat=True).distinct())
    if year not in years:
        years.append(year)
    years.sort(reverse=True)
    return render(request, 'school/incoming_list.html', {
        'letters': letters,
        'year': year,
        'years': years,
        'today': date.today(),
    })


@login_required
def incoming_delete(request, letter_id):
    if not has_perm(request.user, 'incoming', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    letter = get_object_or_404(IncomingLetter, id=letter_id)
    year = letter.academic_year
    letter.delete()
    messages.success(request, 'تم حذف الوارد')
    return redirect(f'{reverse("incoming_list")}?year={year}')


# ─── Secretary: Outgoing (سجل الصادر) ────────────────────────────────────────

@login_required
def outgoing_list(request):
    if not has_perm(request.user, 'outgoing', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST' and has_perm(request.user, 'outgoing', 'add'):
        book_number = request.POST.get('book_number', '').strip()
        if not book_number:
            messages.error(request, 'رقم الكتاب مطلوب')
        else:
            OutgoingLetter.objects.create(
                book_number=book_number,
                book_date=request.POST.get('book_date', '') or date.today(),
                department=request.POST.get('department', '').strip(),
                subject=request.POST.get('subject', '').strip(),
                issuing_entity=request.POST.get('issuing_entity', '').strip(),
                created_by=request.user,
            )
            messages.success(request, 'تم إضافة الصادر بنجاح')
        return redirect('outgoing_list')
    letters = OutgoingLetter.objects.all().order_by('book_date', 'book_number')
    return render(request, 'school/outgoing_list.html', {
        'letters': letters,
        'today': date.today(),
    })


@login_required
def outgoing_delete(request, letter_id):
    if not has_perm(request.user, 'outgoing', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    letter = get_object_or_404(OutgoingLetter, id=letter_id)
    letter.delete()
    messages.success(request, 'تم حذف الصادر')
    return redirect('outgoing_list')


# ─── Teacher Monthly Follow-up (متابعة المعلمين) ─────────────────────────────

MONTHS_AR = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']


def parse_month(value):
    """يقبل 'YYYY-MM' ويرجع (year, month) أو الشهر الحالي إذا لم يكن صالحاً."""
    try:
        if value and '-' in value:
            y, m = value.split('-')
            y, m = int(y), int(m)
            if 1 <= m <= 12:
                return y, m
    except (ValueError, TypeError):
        pass
    today = date.today()
    return today.year, today.month


def month_label(year, month):
    return f'{MONTHS_AR[month - 1]} {year}'


@login_required
def teacher_followups(request):
    if not has_perm(request.user, 'teacher_followup', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST' and has_perm(request.user, 'teacher_followup', 'add'):
        teacher_id = request.POST.get('teacher_id', '')
        teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None
        if not teacher:
            messages.error(request, 'اختر معلماً')
        else:
            TeacherFollowup.objects.create(
                teacher=teacher,
                follow_date=request.POST.get('follow_date', '') or date.today(),
                prep_done='prep_done' in request.POST,
                prep_notes=request.POST.get('prep_notes', '').strip(),
                marks_done='marks_done' in request.POST,
                marks_notes=request.POST.get('marks_notes', '').strip(),
                follow_done='follow_done' in request.POST,
                follow_notes=request.POST.get('follow_notes', '').strip(),
                plans_done='plans_done' in request.POST,
                plans_notes=request.POST.get('plans_notes', '').strip(),
                absence_done='absence_done' in request.POST,
                absence_notes=request.POST.get('absence_notes', '').strip(),
                general_notes=request.POST.get('general_notes', '').strip(),
                created_by=request.user,
            )
            messages.success(request, 'تم تسجيل متابعة المعلم بنجاح')
        return redirect('teacher_followups')
    month = request.GET.get('month', '')
    is_all = month == 'all'
    if is_all:
        followups = TeacherFollowup.objects.select_related('teacher', 'created_by').order_by('teacher__full_name', '-follow_date')
        year, mon, month_value, month_label_text = None, None, 'all', 'جميع الشهور'
    else:
        year, mon = parse_month(month)
        followups = TeacherFollowup.objects.filter(follow_date__year=year, follow_date__month=mon).select_related('teacher', 'created_by').order_by('teacher__full_name', '-follow_date')
        month_value, month_label_text = f'{year}-{mon:02d}', month_label(year, mon)
    teachers = Teacher.objects.all().order_by('full_name')
    return render(request, 'school/teacher_followups.html', {
        'followups': followups,
        'teachers': teachers,
        'today': date.today(),
        'year': year,
        'mon': mon,
        'month_value': month_value,
        'month_label': month_label_text,
        'is_all': is_all,
        'months_ar': MONTHS_AR,
    })


@login_required
def teacher_followup_delete(request, followup_id):
    if not has_perm(request.user, 'teacher_followup', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    f = get_object_or_404(TeacherFollowup, id=followup_id)
    f.delete()
    messages.success(request, 'تم حذف متابعة المعلم')
    return redirect('teacher_followups')


@login_required
def teacher_followup_report(request):
    if not has_perm(request.user, 'teacher_followup', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    month = request.GET.get('month', '')
    is_all = month == 'all'
    if is_all:
        followups = TeacherFollowup.objects.select_related('teacher').order_by('teacher__full_name', '-follow_date')
        year, mon, month_value, month_label_text = None, None, 'all', 'جميع الشهور'
    else:
        year, mon = parse_month(month)
        followups = TeacherFollowup.objects.filter(follow_date__year=year, follow_date__month=mon).select_related('teacher').order_by('teacher__full_name', '-follow_date')
        month_value, month_label_text = f'{year}-{mon:02d}', month_label(year, mon)
    info = SchoolInfo.objects.first()
    return render(request, 'school/teacher_followup_report.html', {
        'followups': followups,
        'year': year,
        'mon': mon,
        'month_value': month_value,
        'month_label': month_label_text,
        'is_all': is_all,
        'info': info,
    })


@login_required
def teacher_followup_missing(request):
    if not has_perm(request.user, 'teacher_followup', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    month = request.GET.get('month', '')
    is_all = month == 'all'
    if is_all:
        followed_ids = TeacherFollowup.objects.values_list('teacher_id', flat=True)
        year, mon, month_value, month_label_text = None, None, 'all', 'جميع الشهور'
    else:
        year, mon = parse_month(month)
        followed_ids = TeacherFollowup.objects.filter(follow_date__year=year, follow_date__month=mon).values_list('teacher_id', flat=True)
        month_value, month_label_text = f'{year}-{mon:02d}', month_label(year, mon)
    missing = Teacher.objects.exclude(id__in=followed_ids).order_by('full_name')
    info = SchoolInfo.objects.first()
    return render(request, 'school/teacher_followup_missing.html', {
        'missing': missing,
        'year': year,
        'mon': mon,
        'month_value': month_value,
        'month_label': month_label_text,
        'is_all': is_all,
        'total_teachers': Teacher.objects.count(),
        'info': info,
    })


@login_required
def reciprocal_visit_list(request):
    if not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST' and has_perm(request.user, 'reciprocal_visits', 'add'):
        visitor_id = request.POST.get('visitor_id', '')
        host_id = request.POST.get('host_id', '')
        visitor = get_object_or_404(Teacher, id=visitor_id) if visitor_id else None
        host = get_object_or_404(Teacher, id=host_id) if host_id else None
        if not visitor or not host:
            messages.error(request, 'اختر المعلم الزائر والمعلم المزار')
        elif visitor == host:
            messages.error(request, 'يجب أن يختلف الزائر عن المزار')
        else:
            period = int(request.POST.get('period', 1) or 1)
            class_id = request.POST.get('class_id', '')
            student_class = get_object_or_404(Class, id=class_id) if class_id else None
            ReciprocalVisit.objects.create(
                visitor=visitor,
                host=host,
                visit_date=request.POST.get('visit_date', '') or date.today(),
                student_class=student_class,
                period=period,
                created_by=request.user,
            )
            messages.success(request, 'تم تسجيل الزيارة التبادلية بنجاح')
        return redirect('reciprocal_visit_list')
    visits = ReciprocalVisit.objects.select_related('visitor', 'host', 'student_class').order_by('-visit_date', '-created_at')
    teachers = Teacher.objects.all().order_by('full_name')
    classes = Class.objects.all().order_by('name')
    info = SchoolInfo.objects.first()
    return render(request, 'school/reciprocal_visits.html', {
        'visits': visits,
        'teachers': teachers,
        'classes': classes,
        'today': date.today(),
        'periods': range(1, 8),
        'info': info,
        'can_add': has_perm(request.user, 'reciprocal_visits', 'add'),
        'can_delete': has_perm(request.user, 'reciprocal_visits', 'delete'),
    })


@login_required
def reciprocal_visit_print(request, visit_id):
    if not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(ReciprocalVisit.objects.select_related('visitor', 'host', 'student_class'), id=visit_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/reciprocal_visit_print.html', {
        'visit': visit,
        'info': info,
        'today': date.today(),
    })


@login_required
def reciprocal_visit_report(request, visit_id):
    if not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(ReciprocalVisit.objects.select_related('visitor', 'host', 'student_class', 'filled_by'), id=visit_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/reciprocal_visit_report.html', {
        'visit': visit,
        'info': info,
        'today': date.today(),
    })


@login_required
def reciprocal_visits_room_report(request):
    if not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    pending = ReciprocalVisit.objects.filter(completed=False).select_related('visitor', 'host', 'student_class').order_by('visit_date', '-created_at')
    done = ReciprocalVisit.objects.filter(completed=True).select_related('visitor', 'host', 'student_class').order_by('-visit_date', '-created_at')
    info = SchoolInfo.objects.first()
    return render(request, 'school/reciprocal_visits_room_report.html', {
        'pending': pending,
        'done': done,
        'info': info,
        'today': date.today(),
    })


@login_required
def reciprocal_visit_feedback(request, visit_id):
    if not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(ReciprocalVisit.objects.select_related('visitor', 'host', 'student_class'), id=visit_id)
    teacher = getattr(request.user, 'teacher_profile', None)
    is_manager = has_perm(request.user, 'reciprocal_visits', 'add')
    if teacher and teacher != visit.visitor:
        messages.error(request, 'هذه الزيارة ليست من زياراتك')
        return redirect('teacher_visits')
    if request.method == 'POST' and (teacher == visit.visitor or is_manager):
        positive = request.POST.get('positive_points', '').strip()
        development = request.POST.get('development_points', '').strip()
        agreement = request.POST.get('agreement', '').strip()
        visit.positive_points = positive
        visit.development_points = development
        visit.agreement = agreement
        visit.completed = bool(positive or development or agreement)
        visit.filled_by = teacher or visit.filled_by
        visit.save()
        messages.success(request, 'تم حفظ ملاحظات الزيارة التبادلية')
        if teacher:
            return redirect('teacher_visits')
        return redirect('reciprocal_visit_list')
    return render(request, 'school/reciprocal_visit_feedback.html', {
        'visit': visit,
        'is_manager': is_manager,
    })


@login_required
def reciprocal_visit_delete(request, visit_id):
    if not has_perm(request.user, 'reciprocal_visits', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    visit = get_object_or_404(ReciprocalVisit, id=visit_id)
    visit.delete()
    messages.success(request, 'تم حذف الزيارة التبادلية')
    return redirect('reciprocal_visit_list')


@login_required
def teacher_visits(request):
    teacher = getattr(request.user, 'teacher_profile', None)
    if not teacher or not has_perm(request.user, 'reciprocal_visits', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    inspection = teacher.inspection_visits.order_by('-visit_date').first()
    supervisor = teacher.supervisor_visits.order_by('-visit_date').first()
    today = date.today()
    visit_entries = teacher.visit_program_entries.filter(visit_date__gte=today).order_by('visit_date', '-created_at')
    assigned = ReciprocalVisit.objects.filter(visitor=teacher).select_related('host', 'student_class').order_by('-visit_date', '-created_at')
    assigned = [v for v in assigned if v.completed or v.visit_date >= today]
    guest = ReciprocalVisit.objects.filter(host=teacher).select_related('visitor', 'student_class').order_by('-visit_date', '-created_at')
    info = SchoolInfo.objects.first()
    return render(request, 'school/teacher_visits.html', {
        'teacher': teacher,
        'inspection': inspection,
        'supervisor': supervisor,
        'visit_entries': visit_entries,
        'assigned': assigned,
        'guest': guest,
        'info': info,
        'today': today,
    })


@login_required
def no_objection_list(request):
    if not has_perm(request.user, 'no_objection', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    if request.method == 'POST' and has_perm(request.user, 'no_objection', 'add'):
        student_name = request.POST.get('student_name', '').strip()
        student_class = request.POST.get('student_class', '').strip()
        sending_school = request.POST.get('sending_school', '').strip()
        if not student_name or not student_class or not sending_school:
            messages.error(request, 'يرجى تعبئة جميع الحقول')
        else:
            NoObjection.objects.create(
                student_name=student_name,
                student_class=student_class,
                sending_school=sending_school,
                created_by=request.user,
            )
            messages.success(request, 'تم تسجيل لا مانع بنجاح')
        return redirect('no_objection_list')
    objects = NoObjection.objects.select_related('created_by').order_by('-created_at')
    info = SchoolInfo.objects.first()
    return render(request, 'school/no_objection_list.html', {
        'objects': objects,
        'info': info,
        'can_add': has_perm(request.user, 'no_objection', 'add'),
        'can_delete': has_perm(request.user, 'no_objection', 'delete'),
    })


@login_required
def no_objection_print(request, obj_id):
    if not has_perm(request.user, 'no_objection', 'view'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    obj = get_object_or_404(NoObjection, id=obj_id)
    info = SchoolInfo.objects.first()
    return render(request, 'school/no_objection_print.html', {
        'obj': obj,
        'info': info,
        'today': date.today(),
    })


@login_required
def no_objection_delete(request, obj_id):
    if not has_perm(request.user, 'no_objection', 'delete'):
        messages.error(request, 'ليس لديك صلاحية')
        return redirect('dashboard')
    obj = get_object_or_404(NoObjection, id=obj_id)
    obj.delete()
    messages.success(request, 'تم حذف لا مانع')
    return redirect('no_objection_list')

# ===================== نظام البرامج الأسبوعية =====================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
import json
from datetime import date, datetime
from .models import (SchedulePlan, TeachingLoad, TeacherAvailability,
                     ScheduleConstraint, FixedLesson, ScheduleEntry,
                     Teacher, Class, Subject)
from .scheduling_engine import generate_schedule, evaluate_plan
from .constraint_nlp import parse_constraint_text

DEFAULT_DAYS = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']


def _sch_perm(request, action):
    if not has_perm(request.user, 'schedule', action):
        messages.error(request, 'ليس لديك صلاحية كافية لهذا القسم.')
        return False
    return True


def _subject_colors(subjects):
    colors = {}
    n = max(1, len(subjects))
    for i, s in enumerate(subjects):
        hue = int(360 * i / n)
        colors[s.id] = 'hsl(%d,60%%,52%%)' % hue
    return colors


@login_required
def schedule_plan_list(request):
    if not _sch_perm(request, 'view'):
        return redirect('dashboard')
    plans = SchedulePlan.objects.all()
    return render(request, 'school/schedule_plans.html', {'plans': plans})


@login_required
def schedule_plan_detail(request, plan_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    return render(request, 'school/schedule_plan_detail.html', {'plan': plan})


@login_required
def schedule_plan_settings(request, plan_id=None):
    if plan_id:
        if not _sch_perm(request, 'edit'):
            return redirect('schedule_plan_list')
        plan = get_object_or_404(SchedulePlan, id=plan_id)
    else:
        if not _sch_perm(request, 'add'):
            return redirect('schedule_plan_list')
        plan = None

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        year = request.POST.get('academic_year', '').strip()
        sem = request.POST.get('semester', '').strip()
        if not name:
            messages.error(request, 'الرجاء إدخال اسم البرنامج.')
            return redirect(request.path)
        days = []
        day_names = request.POST.getlist('day_name')
        day_active = request.POST.getlist('day_active')
        for i, dn in enumerate(day_names):
            days.append({'idx': i, 'name': dn.strip() or DEFAULT_DAYS[i % 7],
                         'active': str(i) in day_active})
        periods = []
        per_names = request.POST.getlist('period_name')
        per_start = request.POST.getlist('period_start')
        per_end = request.POST.getlist('period_end')
        per_active = request.POST.getlist('period_active')
        for i, pn in enumerate(per_names):
            periods.append({'idx': i + 1, 'name': pn.strip() or ('الحصة %d' % (i + 1)),
                            'start': (per_start[i] if i < len(per_start) else ''),
                            'end': (per_end[i] if i < len(per_end) else ''),
                            'active': str(i) in per_active})
        if plan is None:
            plan = SchedulePlan.objects.create(name=name, academic_year=year,
                                               semester=sem, created_by=request.user)
        plan.name = name
        plan.academic_year = year
        plan.semester = sem
        plan.days = days
        plan.periods = periods
        plan.save()
        messages.success(request, 'تم حفظ إعدادات البرنامج.')
        return redirect('schedule_plan_detail', plan_id=plan.id)

    if plan is None:
        days = [{'idx': i, 'name': DEFAULT_DAYS[i], 'active': True} for i in range(5)]
        periods = [{'idx': i + 1, 'name': 'الحصة %d' % (i + 1), 'start': '', 'end': '', 'active': True} for i in range(7)]
    else:
        days = plan.days or []
        periods = plan.periods or []
    return render(request, 'school/schedule_settings.html',
                  {'plan': plan, 'days': days, 'periods': periods})


@login_required
def teaching_loads(request, plan_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'split':
            lid = request.POST.get('load_id')
            tl = plan.teaching_loads.filter(id=lid).first()
            if tl:
                n = tl.weekly_periods
                first = (n + 1) // 2
                second = n - first
                tl.delete()
                if first:
                    TeachingLoad.objects.create(plan=plan, teacher=tl.teacher, subject=tl.subject,
                                                student_class=tl.student_class, weekly_periods=first,
                                                semester='الأول')
                if second:
                    TeachingLoad.objects.create(plan=plan, teacher=tl.teacher, subject=tl.subject,
                                                student_class=tl.student_class, weekly_periods=second,
                                                semester='الثاني')
                messages.success(request, 'تم تقسيم النصاب على الفصلين الدراسيين.')
            return redirect('teaching_loads', plan_id=plan.id)
        if action == 'delete':
            lid = request.POST.get('load_id')
            plan.teaching_loads.filter(id=lid).delete()
            messages.success(request, 'تم حذف النصاب.')
            return redirect('teaching_loads', plan_id=plan.id)
        tid = request.POST.get('teacher')
        sid = request.POST.get('subject')
        cid = request.POST.get('student_class')
        wp = int(request.POST.get('weekly_periods', 1) or 1)
        sem = request.POST.get('semester', '')
        lid = request.POST.get('load_id')
        if tid and sid and cid:
            if lid:
                tl = plan.teaching_loads.filter(id=lid).first()
                if tl:
                    tl.teacher_id = tid
                    tl.subject_id = sid
                    tl.student_class_id = cid
                    tl.weekly_periods = wp
                    tl.semester = sem
                    tl.save()
                    messages.success(request, 'تم تحديث النصاب.')
            else:
                TeachingLoad.objects.update_or_create(
                    plan=plan, teacher_id=tid, subject_id=sid, student_class_id=cid,
                    defaults={'weekly_periods': wp, 'semester': sem})
                messages.success(request, 'تم حفظ النصاب.')
        return redirect('teaching_loads', plan_id=plan.id)
    loads = plan.teaching_loads.select_related('teacher', 'subject', 'student_class').all()
    agg = {}
    order = []
    for l in loads:
        k = l.teacher_id
        if k not in agg:
            agg[k] = {'teacher': l.teacher, 'total': 0, 'items': []}
            order.append(k)
        agg[k]['total'] += l.weekly_periods
        agg[k]['items'].append(l)
    teacher_data = [agg[k] for k in order]
    total_periods = sum(t['total'] for t in teacher_data)
    class_agg = {}
    class_order = []
    for l in loads:
        k = l.student_class_id
        if k not in class_agg:
            class_agg[k] = {'cls': l.student_class, 'total': 0, 'items': []}
            class_order.append(k)
        class_agg[k]['total'] += l.weekly_periods
        class_agg[k]['items'].append(l)
    class_summary = [class_agg[k] for k in class_order]
    class_summary.sort(key=lambda x: x['cls'].name)
    teachers = Teacher.objects.all().order_by('full_name')
    subjects = Subject.objects.all().order_by('name')
    classes = Class.objects.all().order_by('name')
    return render(request, 'school/teaching_loads.html',
                  {'plan': plan, 'loads': loads, 'teacher_data': teacher_data,
                   'class_summary': class_summary, 'total_periods': total_periods,
                   'teachers': teachers, 'subjects': subjects, 'classes': classes})


@login_required
def availability_grid(request, plan_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    teachers = Teacher.objects.all().order_by('full_name')
    days = plan.active_days
    periods = plan.active_periods
    existing = {}
    for a in plan.availabilities.all():
        existing[(a.teacher_id, a.day, a.period)] = a.available
    teacher_data = []
    for t in teachers:
        rows = []
        for d in days:
            cells = []
            for p in periods:
                cells.append({'day': d['name'], 'period': p['idx'],
                              'avail': existing.get((t.id, d['name'], p['idx']), True)})
            rows.append({'day': d['name'], 'cells': cells})
        teacher_data.append({'teacher': t, 'rows': rows})
    if request.method == 'POST':
        plan.availabilities.all().delete()
        bulk = []
        for t in teachers:
            for d in days:
                for p in periods:
                    key = 'avail_%d_%s_%s' % (t.id, d['name'], p['idx'])
                    bulk.append(TeacherAvailability(
                        plan=plan, teacher=t, day=d['name'], period=p['idx'],
                        available=(key in request.POST)))
        TeacherAvailability.objects.bulk_create(bulk)
        messages.success(request, 'تم حفظ تفريغ المعلمين.')
        return redirect('availability_grid', plan_id=plan.id)
    return render(request, 'school/availability_grid.html',
                  {'plan': plan, 'teacher_data': teacher_data, 'periods': periods})


@login_required
def schedule_constraints(request, plan_id):
    if not _sch_perm(request, 'manage_constraints'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            code = request.POST.get('code', '').strip()
            label = request.POST.get('label', '').strip()
            params = {}
            if code == 'spread_subject':
                params = {'max_per_day': int(request.POST.get('max_per_day', 1) or 1)}
            elif code == 'max_consecutive_gap':
                params = {'max_gap': int(request.POST.get('max_gap', 1) or 1)}
            elif code == 'period_repeat':
                params = {'period': int(request.POST.get('period', 1) or 1),
                          'max_days': int(request.POST.get('max_days', 1) or 1)}
            scope = request.POST.get('scope', 'all')
            obj = ScheduleConstraint.objects.create(
                plan=plan, type=request.POST.get('type', 'hard'),
                code=code, label=label, enabled=True, scope=scope,
                weight=float(request.POST.get('weight', 1) or 1),
                params=params)
            if scope == 'teachers':
                obj.teachers.set([int(x) for x in request.POST.getlist('teachers') if x])
            elif scope == 'classes':
                obj.classes.set([int(x) for x in request.POST.getlist('classes') if x])
            messages.success(request, 'تمت إضافة الشرط.')
        elif action == 'command':
            parsed = parse_constraint_text(request.POST.get('command', ''), plan)
            if not parsed:
                messages.error(request, 'لم أفهم الشرط. جرّب صياغة أوضح مثل: «المعلم أحمد لا يزيد عن 3 حصص متتالية».')
            else:
                obj = ScheduleConstraint.objects.create(
                    plan=plan, type=parsed['type'], code=parsed['code'],
                    label=parsed['label'], enabled=True, scope=parsed['scope'],
                    weight=parsed['weight'], params=parsed['params'])
                if parsed['scope'] == 'teachers':
                    obj.teachers.set(parsed['teacher_ids'])
                elif parsed['scope'] == 'classes':
                    obj.classes.set(parsed['class_ids'])
                messages.success(request,
                    'تم فهم الشرط وتطبيقه: %s (النطاق: %s).' % (obj.label, obj.get_scope_display()))
        elif action == 'edit':
            obj = get_object_or_404(ScheduleConstraint, id=request.POST.get('cid'), plan=plan)
            code = request.POST.get('code', '').strip()
            label = request.POST.get('label', '').strip()
            params = {}
            if code == 'spread_subject':
                params = {'max_per_day': int(request.POST.get('max_per_day', 1) or 1)}
            elif code == 'max_consecutive_gap':
                params = {'max_gap': int(request.POST.get('max_gap', 1) or 1)}
            elif code == 'period_repeat':
                params = {'period': int(request.POST.get('period', 1) or 1),
                          'max_days': int(request.POST.get('max_days', 1) or 1)}
            obj.type = request.POST.get('type', obj.type)
            obj.code = code
            obj.label = label
            obj.weight = float(request.POST.get('weight', obj.weight) or obj.weight)
            obj.scope = request.POST.get('scope', obj.scope)
            obj.params = params
            obj.save()
            if obj.scope == 'teachers':
                obj.teachers.set([int(x) for x in request.POST.getlist('teachers') if x])
                obj.classes.clear()
            elif obj.scope == 'classes':
                obj.classes.set([int(x) for x in request.POST.getlist('classes') if x])
                obj.teachers.clear()
            else:
                obj.teachers.clear()
                obj.classes.clear()
            messages.success(request, 'تم تعديل الشرط.')
        elif action == 'toggle':
            c = get_object_or_404(ScheduleConstraint, id=request.POST.get('cid'))
            c.enabled = not c.enabled
            c.save()
        elif action == 'delete':
            get_object_or_404(ScheduleConstraint, id=request.POST.get('cid')).delete()
        return redirect('schedule_constraints', plan_id=plan.id)
    constraints = plan.constraints.all().order_by('type', 'code')
    tl_qs = TeachingLoad.objects.filter(plan=plan)
    teacher_ids = tl_qs.values_list('teacher_id', flat=True)
    class_ids = tl_qs.values_list('student_class_id', flat=True)
    teachers = Teacher.objects.filter(id__in=teacher_ids).order_by('full_name')
    classes = Class.objects.filter(id__in=class_ids).order_by('name')
    if not teachers.exists():
        teachers = Teacher.objects.all().order_by('full_name')
    if not classes.exists():
        classes = Class.objects.all().order_by('name')
    return render(request, 'school/schedule_constraints.html',
                  {'plan': plan, 'constraints': constraints,
                   'teachers': teachers, 'classes': classes})


@login_required
def fixed_lessons(request, plan_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        if action == 'delete':
            fid = request.POST.get('fixed_id')
            plan.fixed_lessons.filter(id=fid).delete()
            messages.success(request, 'تم حذف التثبيت.')
            return redirect('fixed_lessons', plan_id=plan.id)
        day = request.POST.get('day')
        period = int(request.POST.get('period') or 0)
        tid = request.POST.get('teacher')
        sid = request.POST.get('subject')
        cid = request.POST.get('student_class')
        fid = request.POST.get('fixed_id')
        if day and period and tid and sid and cid:
            if fid:
                plan.fixed_lessons.filter(id=fid).delete()
            FixedLesson.objects.update_or_create(
                plan=plan, day=day, period=period,
                defaults={'teacher_id': tid, 'subject_id': sid, 'student_class_id': cid})
            messages.success(request, 'تم تثبيت الحصة.')
        return redirect('fixed_lessons', plan_id=plan.id)
    fixed = plan.fixed_lessons.select_related('teacher', 'subject', 'student_class').all()
    days = plan.active_days
    periods = plan.active_periods
    teachers = Teacher.objects.all().order_by('full_name')
    subjects = Subject.objects.all().order_by('name')
    classes = Class.objects.all().order_by('name')
    return render(request, 'school/fixed_lessons.html',
                  {'plan': plan, 'fixed': fixed, 'days': days, 'periods': periods,
                   'teachers': teachers, 'subjects': subjects, 'classes': classes})


@login_required
@transaction.atomic
def schedule_generate(request, plan_id):
    if not _sch_perm(request, 'generate'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    has_loads = plan.teaching_loads.exists()
    if request.method == 'POST':
        if not has_loads:
            messages.error(request, 'لا توجد أنصبة مسجلة بعد. سجّل الأنصبة أولًا.')
            return redirect('schedule_generate', plan_id=plan.id)
        try:
            result = generate_schedule(plan)
        except Exception as ex:
            messages.error(request, 'خطأ أثناء التوليد: %s' % ex)
            return redirect('schedule_generate', plan_id=plan.id)
        ScheduleEntry.objects.filter(plan=plan).delete()
        color_map = _subject_colors(Subject.objects.all())
        entries = []
        for e in result['entries']:
            entries.append(ScheduleEntry(
                plan=plan, day=e['day'], period=e['period'], teacher_id=e['teacher'],
                subject_id=e['subject'], student_class_id=e['class'],
                fixed=e['fixed'], color=color_map.get(e['subject'], '')))
        ScheduleEntry.objects.bulk_create(entries)
        plan.hard_score = result['hard_score']
        plan.soft_score = result['soft_score']
        plan.generated_at = timezone.now()
        plan.status = 'active'
        plan.save()
        if result['unscheduled']:
            tmap = {t.id: t.full_name for t in Teacher.objects.all()}
            smap = {s.id: s.name for s in Subject.objects.all()}
            cmap = {c.id: c.name for c in Class.objects.all()}
            details = []
            for u in result['unscheduled']:
                details.append('%s — %s (%s): %s' % (tmap.get(u['teacher'], u['teacher']),
                                                 smap.get(u['subject'], u['subject']),
                                                 cmap.get(u['class'], u['class']),
                                                 u.get('reason', '')))
            messages.warning(request, 'تعذّر جدولة %d حصة (أنصبة زائدة أو تعارضات صلبة): %s' % (
                len(details), ' | '.join(details)))
        elif result['hard_score'] < 100:
            messages.warning(request, 'تم الإنشاء لكن توجد تعارضات صلبة (صلب %s%%). راجع التقرير أدناه.' % result['hard_score'])
        else:
            messages.success(request, 'تم إنشاء البرنامج (صلب %s%% / ناعم %s%%).' % (
                result['hard_score'], result['soft_score']))
        return redirect('schedule_grid', plan_id=plan.id)
    return render(request, 'school/schedule_generate.html', {'plan': plan, 'has_loads': has_loads})


def _conflict_texts(plan):
    try:
        conflicts = evaluate_plan(plan)['conflicts']
    except Exception:
        return []
    teachers = {t.id: t.full_name for t in Teacher.objects.all()}
    classes = {c.id: c.name for c in Class.objects.all()}
    subjects = {s.id: s.name for s in Subject.objects.all()}
    out = []
    for c in conflicts:
        t = c.get('teacher'); cl = c.get('student_class'); s = c.get('subject')
        day = c.get('day'); period = c.get('period')
        tp = 'يوم %s حصة %s' % (day, period)
        if c['type'] == 'teacher_double':
            out.append('تعارض: المعلم %s مسجّل مرتين (%s).' % (teachers.get(t, t), tp))
        elif c['type'] == 'class_double':
            out.append('تعارض: الصف %s مسجّل مرتين (%s).' % (classes.get(cl, cl), tp))
        elif c['type'] == 'availability':
            out.append('تعارض توفّر: المعلم %s غير متاح (%s).' % (teachers.get(t, t), tp))
        elif c['type'] == 'fixed_violation':
            out.append('مخالفة حصة مثبّتة (%s).' % tp)
        elif c['type'] == 'max_per_day':
            who = teachers.get(t, t) or classes.get(cl, cl)
            out.append('تجاوز أقصى حصص في اليوم لـ %s (%s).' % (who, tp))
        elif c['type'] == 'spread':
            out.append('تكرار المادة %s للمعلم %s أكثر من المسموح يوم %s.' % (subjects.get(s, s), teachers.get(t, t), day))
        elif c['type'] == 'gap':
            out.append('تجاوز أقصى فراغات متتالية للمعلم %s يوم %s.' % (teachers.get(t, t), day))
        elif c['type'] == 'period_repeat':
            out.append('تكرار الحصة %s للمعلم %s أكثر من الأيام المسموحة.' % (period, teachers.get(t, t)))
        else:
            out.append('تعارض غير مصنّف (%s).' % tp)
    return out


def _grid_context(plan, teacher=None, student_class=None):
    days = plan.active_days
    periods = plan.active_periods
    color_map = _subject_colors(Subject.objects.all())
    q = plan.entries.select_related('teacher', 'subject', 'student_class').all()
    if teacher is not None:
        q = q.filter(teacher=teacher)
    if student_class is not None:
        q = q.filter(student_class=student_class)
    cells = defaultdict(list)
    for e in q:
        cells[(e.day, e.period)].append(e)
    conflict_keys = set()
    try:
        for c in evaluate_plan(plan)['conflicts']:
            if 'day' in c and 'period' in c:
                conflict_keys.add((c['day'], c['period'], c.get('teacher'), c.get('student_class')))
    except Exception:
        pass
    conflict_texts = _conflict_texts(plan)
    for lst in cells.values():
        for cell in lst:
            if (cell.day, cell.period, cell.teacher_id, cell.student_class_id) in conflict_keys:
                cell.conflict_flag = True
    grid_data = []
    for p in periods:
        row = [{'day': d['name'], 'period': p['idx'], 'cell': cells.get((d['name'], p['idx']), [])}
               for d in days]
        grid_data.append((p, row))
    conflict_cells = {(k[0], k[1]) for k in conflict_keys}
    return {'plan': plan, 'days': days, 'periods': periods,
            'cells': cells, 'grid_data': grid_data, 'color_map': color_map,
            'conflict_cells': conflict_cells, 'conflict_texts': conflict_texts}


@login_required
def schedule_grid(request, plan_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    ctx = _grid_context(plan)
    ctx['plan'] = plan
    return render(request, 'school/schedule_grid.html', ctx)


@login_required
def schedule_teacher(request, plan_id, teacher_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    teacher = get_object_or_404(Teacher, id=teacher_id)
    ctx = _grid_context(plan, teacher=teacher)
    ctx['teacher'] = teacher
    return render(request, 'school/schedule_teacher.html', ctx)


@login_required
def schedule_class(request, plan_id, class_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    cls = get_object_or_404(Class, id=class_id)
    ctx = _grid_context(plan, student_class=cls)
    ctx['class'] = cls
    return render(request, 'school/schedule_class.html', ctx)


@login_required
def availability_report(request, plan_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    teachers = Teacher.objects.all().order_by('full_name')
    days = plan.active_days
    periods = plan.active_periods
    existing = {}
    for a in plan.availabilities.all():
        existing[(a.teacher_id, a.day, a.period)] = a.available
    placed = {}
    for e in plan.entries.all():
        placed[(e.teacher_id, e.day, e.period)] = True
    teacher_data = []
    for t in teachers:
        rows = []
        for d in days:
            cells = []
            for p in periods:
                cells.append({'avail': existing.get((t.id, d['name'], p['idx'])),
                              'placed': placed.get((t.id, d['name'], p['idx'], False))})
            rows.append({'day': d['name'], 'cells': cells})
        teacher_data.append({'teacher': t, 'rows': rows})
    return render(request, 'school/availability_report.html',
                  {'plan': plan, 'teacher_data': teacher_data, 'periods': periods})


@login_required
def schedule_plan_copy(request, plan_id):
    if not _sch_perm(request, 'add'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        np = SchedulePlan.objects.create(
            name=plan.name + ' (نسخة)', academic_year=plan.academic_year,
            semester=plan.semester, days=plan.days, periods=plan.periods,
            created_by=request.user, parent=plan)
        for tl in plan.teaching_loads.all():
            TeachingLoad.objects.create(plan=np, teacher=tl.teacher, subject=tl.subject,
                                        student_class=tl.student_class, weekly_periods=tl.weekly_periods)
        for a in plan.availabilities.all():
            TeacherAvailability.objects.create(plan=np, teacher=a.teacher, day=a.day,
                                               period=a.period, available=a.available)
        for c in plan.constraints.all():
            ScheduleConstraint.objects.create(plan=np, type=c.type, code=c.code,
                                               label=c.label, enabled=c.enabled, weight=c.weight, params=c.params)
        for f in plan.fixed_lessons.all():
            FixedLesson.objects.create(plan=np, day=f.day, period=f.period, teacher=f.teacher,
                                       subject=f.subject, student_class=f.student_class)
        messages.success(request, 'تم نسخ البرنامج.')
        return redirect('schedule_plan_detail', plan_id=np.id)
    return redirect('schedule_plan_detail', plan_id=plan.id)


@login_required
def schedule_plan_activate(request, plan_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    SchedulePlan.objects.filter(status='active').update(status='archived')
    plan.status = 'active'
    plan.save()
    messages.success(request, 'تم تفعيل البرنامج.')
    return redirect('schedule_plan_list')


@login_required
def schedule_plan_delete(request, plan_id):
    if not _sch_perm(request, 'delete'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        plan.delete()
        messages.success(request, 'تم حذف البرنامج.')
    return redirect('schedule_plan_list')

@login_required
def schedule_export_excel(request, plan_id):
    if not _sch_perm(request, 'export'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    days = plan.active_days
    periods = plan.active_periods
    color_map = _subject_colors(Subject.objects.all())
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='3498DB')
    hdr_font = Font(bold=True, color='FFFFFF')

    entries = {}
    for e in plan.entries.select_related('subject', 'teacher', 'student_class').all():
        entries[(e.day, e.period)] = e

    # ورقة البرنامج المدرسي
    ws = wb.active
    ws.title = 'البرنامج'
    ws.append(['الحصة'] + [d['name'] for d in days])
    for p in periods:
        row = [p['name']]
        for d in days:
            e = entries.get((d['name'], p['idx']))
            row.append(('%s\n%s\n%s' % (e.subject.name, e.teacher.full_name, e.student_class.name)) if e else '')
        ws.append(row)
    for c in range(1, len(days) + 2):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r in range(2, len(periods) + 2):
        for c in range(1, len(days) + 2):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i in range(len(periods) + 1):
        ws.column_dimensions[get_column_letter(1)].width = 14
    for c in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # ورقة لكل معلم
    for teacher in Teacher.objects.all().order_by('full_name'):
        ws2 = wb.create_sheet(name=teacher.full_name[:28] or 'معلم')
        ws2.append(['الحصة'] + [d['name'] for d in days])
        te = {}
        for e in plan.entries.filter(teacher=teacher).select_related('subject', 'student_class'):
            te[(e.day, e.period)] = e
        for p in periods:
            row = [p['name']]
            for d in days:
                e = te.get((d['name'], p['idx']))
                row.append(('%s\n%s' % (e.subject.name, e.student_class.name)) if e else '')
            ws2.append(row)
        for c in range(1, len(days) + 2):
            cell = ws2.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r in range(2, len(periods) + 2):
            for c in range(1, len(days) + 2):
                cell = ws2.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    from urllib.parse import quote
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % quote(plan.name)
    wb.save(response)
    return response


@login_required
def schedule_print_grid(request, plan_id):
    if not _sch_perm(request, 'print'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    ctx = _grid_context(plan)
    ctx['plan'] = plan
    ctx['print_mode'] = True
    return render(request, 'school/schedule_grid.html', ctx)

@login_required
def schedule_teachers(request, plan_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    teachers = Teacher.objects.all().order_by('full_name')
    return render(request, 'school/schedule_picker.html',
                  {'plan': plan, 'items': teachers, 'kind': 'teacher'})


@login_required
def schedule_classes(request, plan_id):
    if not _sch_perm(request, 'view'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    classes = Class.objects.all().order_by('name')
    return render(request, 'school/schedule_picker.html',
                  {'plan': plan, 'items': classes, 'kind': 'class'})


@login_required
def schedule_edit_grid(request, plan_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    entries = plan.entries.select_related('teacher', 'subject', 'student_class').order_by('day', 'period')
    teachers = Teacher.objects.all().order_by('full_name')
    subjects = Subject.objects.all().order_by('name')
    classes = Class.objects.all().order_by('name')
    days = plan.active_days
    periods = plan.active_periods
    if request.method == 'POST':
        delete_ids = set(request.POST.getlist('delete'))
        new_vals = {}
        for e in entries:
            if e.id in delete_ids:
                continue
            new_vals[e.id] = {
                'teacher': int(request.POST.get('teacher_%d' % e.id, e.teacher_id)),
                'subject': int(request.POST.get('subject_%d' % e.id, e.subject_id)),
                'class': int(request.POST.get('class_%d' % e.id, e.student_class_id)),
                'day': request.POST.get('day_%d' % e.id, e.day),
                'period': int(request.POST.get('period_%d' % e.id, e.period)),
            }
        seen_t = {}
        seen_c = {}
        conflict = False
        for eid, v in new_vals.items():
            kt = (v['teacher'], v['day'], v['period'])
            kc = (v['class'], v['day'], v['period'])
            if kt in seen_t or kc in seen_c:
                conflict = True
            seen_t[kt] = eid
            seen_c[kc] = eid
        if conflict:
            messages.error(request, 'يوجد تعارض: حصتان في نفس المعلم أو نفس الصف والحصة. راجع الإدخالات.')
            return redirect('schedule_edit_grid', plan_id=plan.id)
        for e in entries:
            if e.id in delete_ids:
                e.delete()
                continue
            v = new_vals[e.id]
            e.teacher_id = v['teacher']
            e.subject_id = v['subject']
            e.student_class_id = v['class']
            e.day = v['day']
            e.period = v['period']
            e.save()
        res = evaluate_plan(plan)
        plan.hard_score = res['hard_score']
        plan.soft_score = res['soft_score']
        plan.generated_at = timezone.now()
        plan.status = 'active'
        plan.save()
        messages.success(request, 'تم حفظ التعديل اليدوي (صلب %s%% / ناعم %s%%).' % (
            res['hard_score'], res['soft_score']))
        return redirect('schedule_edit_grid', plan_id=plan.id)
    return render(request, 'school/schedule_edit_grid.html',
                  {'plan': plan, 'entries': entries, 'teachers': teachers,
                   'subjects': subjects, 'classes': classes, 'days': days, 'periods': periods})

@login_required
def schedule_entry_move(request, plan_id, entry_id):
    if not _sch_perm(request, 'edit'):
        return redirect('schedule_plan_list')
    plan = get_object_or_404(SchedulePlan, id=plan_id)
    if request.method == 'POST':
        e = get_object_or_404(ScheduleEntry, id=entry_id, plan=plan)
        day = request.POST.get('day')
        period = int(request.POST.get('period', e.period) or e.period)
        if day:
            e.day = day
            e.period = period
            e.save()
        res = evaluate_plan(plan)
        plan.hard_score = res['hard_score']
        plan.soft_score = res['soft_score']
        plan.generated_at = timezone.now()
        plan.status = 'active'
        plan.save()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'hard_score': res['hard_score'],
                                 'soft_score': res['soft_score']})
        messages.success(request, 'تم نقل الحصة.')
    return redirect('schedule_grid', plan_id=plan.id)

